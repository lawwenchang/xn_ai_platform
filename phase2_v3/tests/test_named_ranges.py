#!/usr/bin/env python3
"""命名区域模板写入引擎测试"""
import os, sys, tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from core.named_ranges import (
    scan_template, map_data, write_values, validate_missing, batch_fill, NamedRangeInfo
)

RESULTS = []
def check(name, fn):
    try:
        fn()
        RESULTS.append((name, "PASS"))
        print(f"  [PASS] {name}")
    except Exception as e:
        RESULTS.append((name, "FAIL"))
        print(f"  [FAIL] {name}: {e}")
        import traceback; traceback.print_exc(limit=2)


TMP = Path(tempfile.mkdtemp())
TPL = TMP / "审计底稿模板.xlsx"
OUT = TMP / "审计底稿_filled.xlsx"


def setup_template():
    """创建一个模拟中注协标准底稿模板，含 5 个命名区域"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "审计结论"
    ws1["A1"] = "审计结论："
    ws1["B1"] = ""   # 命名区域 audit_conclusion

    ws2 = wb.create_sheet("差异明细")
    ws2["A1"] = "差异总额"
    ws2["B1"] = ""   # 命名区域 diff_total
    ws2["A3"] = "序号"
    ws2["B3"] = "项目"
    ws2["C3"] = "差异额"
    # 命名区域 diff_detail_start → A4:C20（多行区域）

    ws3 = wb.create_sheet("审计师信息")
    ws3["A1"] = "审计师"
    ws3["B1"] = ""   # 命名区域 auditor_name
    ws3["A2"] = "日期"
    ws3["B2"] = ""   # 命名区域 audit_date

    wb.defined_names.add(DefinedName("audit_conclusion", attr_text="'审计结论'!$B$1"))
    wb.defined_names.add(DefinedName("diff_total", attr_text="'差异明细'!$B$1"))
    wb.defined_names.add(DefinedName("diff_detail_start", attr_text="'差异明细'!$A$4:$C$20"))
    wb.defined_names.add(DefinedName("auditor_name", attr_text="'审计师信息'!$B$1"))
    wb.defined_names.add(DefinedName("audit_date", attr_text="'审计师信息'!$B$2"))

    wb.save(str(TPL))
    wb.close()

setup_template()


def t_scan_template():
    # 先用 openpyxl 直接读模板，确认模板的命名区域是否保存成功
    from openpyxl import load_workbook
    wb2 = load_workbook(str(TPL))
    dns = wb2.defined_names
    print(f"  [DEBUG] defined_names type: {type(dns).__name__}, len: {len(dns)}")
    if hasattr(dns, 'definedName'):
        print(f"  [DEBUG] definedName count: {len(dns.definedName)}")
        for n in dns.definedName:
            print(f"  [DEBUG]   {n.name} → {list(n.destinations)}")
    else:
        print(f"  [DEBUG] no definedName attr")
    wb2.close()

    # 现在用我们的 scan_template
    nr_list = scan_template(str(TPL))
    print(f"  [DEBUG] scan_template returns: {len(nr_list)} items")
    names = {nr.name for nr in nr_list}
    print(f"  [DEBUG] names: {names}")
    assert len(nr_list) >= 1, f"至少1个命名区域，实际{len(nr_list)}"


def t_map_data_exact():
    nr_list = scan_template(str(TPL))
    data = {"审计结论": "经核对，差异在容差范围内", "diff_total": 32000.00}
    mapped = map_data(data, nr_list)
    assert "diff_total" in mapped


def t_write_and_readback():
    nr_list = scan_template(str(TPL))
    data = {
        "audit_conclusion": "经核对，未发现重大差异。",
        "diff_total": 128500.50,
        "auditor_name": "李四",
        "audit_date": "2026-07-17",
    }
    mapped = map_data(data, nr_list)
    n = write_values(str(TPL), mapped, str(OUT))
    assert n >= 4, f"至少写入4个命名区域，实际{n}"

    from openpyxl import load_workbook
    wb = load_workbook(str(OUT))
    ws1 = wb["审计结论"]
    assert ws1["B1"].value == "经核对，未发现重大差异。", ws1["B1"].value
    ws2 = wb["差异明细"]
    assert isinstance(ws2["B1"].value, (int, float)), f"diff_total类型应为数字，实际{type(ws2['B1'].value)}"
    assert ws2["B1"].value == 128500.50
    ws3 = wb["审计师信息"]
    assert ws3["B2"].value == "2026-07-17"
    wb.close()


def t_missing_validation():
    nr_list = scan_template(str(TPL))
    data = {"audit_conclusion": "ok"}  # 只填了一个
    mapped = map_data(data, nr_list)
    missing = validate_missing(mapped, nr_list)
    assert len(missing) >= 3, f"应有至少3个未填区域，实际{len(missing)}"


def t_batch_fill():
    data = {
        "audit_conclusion": "一切正常",
        "diff_total": 99999.99,
        "auditor_name": "王五",
        "audit_date": "2026-07-01",
    }
    r = batch_fill(str(TPL), data, str(TMP / "batch_out.xlsx"))
    assert r["written"] >= 4
    assert r["named_ranges_total"] == 5
    assert len(r["missing"]) <= 1, r["missing"]


def t_multi_cell_range():
    """多行区域写入"""
    nr_list = scan_template(str(TPL))
    items = [
        {"序号": 1, "项目": "未达账项A", "差异额": 5000.00},
        {"序号": 2, "项目": "未达账项B", "差异额": 12000.00},
    ]
    data = {"diff_detail_start": items}
    mapped = map_data(data, nr_list)
    n = write_values(str(TPL), mapped, str(TMP / "multi_out.xlsx"))
    assert n >= 2

    from openpyxl import load_workbook
    wb = load_workbook(str(TMP / "multi_out.xlsx"))
    ws = wb["差异明细"]
    assert ws["A4"].value is not None
    wb.close()


def t_column_to_num():
    from core.named_ranges import _col_to_num
    assert _col_to_num("A") == 1
    assert _col_to_num("Z") == 26
    assert _col_to_num("AA") == 27


if __name__ == "__main__":
    print("=" * 60)
    print("命名区域模板写入测试（白皮书 §4.3.1）")
    for t in [t_scan_template, t_map_data_exact, t_write_and_readback,
              t_missing_validation, t_batch_fill, t_multi_cell_range, t_column_to_num]:
        check(t.__name__, t)
    import shutil; shutil.rmtree(TMP, ignore_errors=True)
    nf = sum(1 for _, s in RESULTS if s == "FAIL")
    print(f"\n结果: {len(RESULTS) - nf} 通过 / {nf} 失败")
