#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
同源勾稽引擎 (cross_reference.py)
==================================
确保报告数字 - 底稿审定数 - 原始数据源三者一致。

白皮书 4.3.1「同源勾稽」的核心实现。
"""

from __future__ import annotations

import json, re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional


@dataclass
class CrossRefEntry:
    """一条勾稽记录"""
    figure_name: str
    report_value: float
    workpaper_value: float
    source_value: float = 0
    tolerance: float = 0.01
    passed: bool = False
    source_path: str = ""
    dag_node: str = ""
    run_id: str = ""
    note: str = ""

    def verify(self) -> bool:
        self.passed = (
            abs(self.report_value - self.workpaper_value) <= self.tolerance
            and abs(self.workpaper_value - self.source_value) <= self.tolerance
        )
        return self.passed

    def to_dict(self) -> dict:
        return {
            "figure": self.figure_name,
            "report_value": self.report_value,
            "workpaper_value": self.workpaper_value,
            "source_value": self.source_value,
            "passed": self.passed,
            "delta_r_w": round(self.report_value - self.workpaper_value, 4),
            "delta_w_s": round(self.workpaper_value - self.source_value, 4),
            "source": self.source_path, "dag_node": self.dag_node,
            "run_id": self.run_id, "note": self.note,
        }


@dataclass
class CrossReferenceReport:
    """勾稽报告"""
    entries: List[CrossRefEntry] = field(default_factory=list)
    generated_at: str = field(default_factory=lambda: datetime.now().isoformat())
    run_id: str = ""
    all_passed: bool = False

    @property
    def passed_count(self) -> int:
        return sum(1 for e in self.entries if e.passed)

    @property
    def failed_count(self) -> int:
        return len(self.entries) - self.passed_count

    def to_dict(self) -> dict:
        return {
            "run_id": self.run_id, "generated_at": self.generated_at,
            "total": len(self.entries), "passed": self.passed_count,
            "failed": self.failed_count, "all_passed": self.all_passed,
            "entries": [e.to_dict() for e in self.entries],
        }

    def to_markdown(self) -> str:
        lines = [
            "# 同源勾稽报告",
            f"Run ID: {self.run_id}",
            f"时间: {self.generated_at}",
            f"总计: {len(self.entries)} | 通过: {self.passed_count} | 失败: {self.failed_count}",
            "",
            "| 数字名称 | 报告值 | 底稿值 | 源值 | 报告-底稿 | 底稿-源 | 状态 |",
            "|---------|--------|--------|------|-----------|----------|------|",
        ]
        for e in self.entries:
            s = "OK" if e.passed else "FAIL"
            lines.append(
                f"| {e.figure_name} | {e.report_value:,.2f} | {e.workpaper_value:,.2f} | "
                f"{e.source_value:,.2f} | {e.report_value-e.workpaper_value:,.2f} | "
                f"{e.workpaper_value-e.source_value:,.2f} | {s} |"
            )


# ═══════════════ 公共 API ═══════════════

def verify_report_against_workpaper(
    report_data: Dict[str, float],
    workpaper_data: Dict[str, float],
    source_data: Optional[Dict[str, float]] = None,
    run_id: str = "",
    tolerance: float = 0.01,
) -> CrossReferenceReport:
    """验证报告数字与底稿数字的一致性"""
    report = CrossReferenceReport(run_id=run_id)
    all_keys = set(report_data.keys()) | set(workpaper_data.keys())

    for key in sorted(all_keys):
        rv = report_data.get(key)
        wv = workpaper_data.get(key)
        sv = source_data.get(key) if source_data else None

        entry = CrossRefEntry(
            figure_name=key,
            report_value=rv if rv is not None else 0,
            workpaper_value=wv if wv is not None else 0,
            source_value=sv if sv is not None else (wv if wv is not None else 0),
            tolerance=tolerance, run_id=run_id,
        )
        if rv is None and wv is not None:
            entry.note = "报告缺少该数字"
        elif rv is not None and wv is None:
            entry.note = "底稿缺少该数字"

        entry.verify()
        report.entries.append(entry)

    report.all_passed = report.failed_count == 0
    return report


def verify_single_figure(
    name: str, report_val: float, workpaper_val: float,
    source_val: float = None, run_id: str = "",
    source_path: str = "", dag_node: str = "", tolerance: float = 0.01,
) -> CrossRefEntry:
    """验证单个数字的勾稽关系"""
    entry = CrossRefEntry(
        figure_name=name, report_value=report_val,
        workpaper_value=workpaper_val,
        source_value=source_val if source_val is not None else workpaper_val,
        tolerance=tolerance, run_id=run_id,
        source_path=source_path, dag_node=dag_node,
    )
    entry.verify()
    return entry


def extract_figures_from_report(report_path: str) -> Dict[str, float]:
    """从Word报告中提取关键数字（格式：名称：金额元）"""
    try:
        from docx import Document
        doc = Document(report_path)
        figures = {}
        for p in doc.paragraphs:
            matches = re.findall(
                r'([\u4e00-\u9fa5（）\w]+)[：:]\s*[\u00a5￥]?\s*([\d,]+\.?\d*)\s*[元万元]',
                p.text
            )
            for name, val in matches:
                try:
                    figures[name.strip()] = float(val.replace(",", ""))
                except ValueError: pass
        return figures
    except Exception: return {}


def extract_figures_from_excel(excel_path: str) -> Dict[str, float]:
    """从Excel底稿中提取数字"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        figures = {}
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                for i, cell in enumerate(row):
                    if isinstance(cell, str) and any(
                        kw in cell for kw in ["合计", "审定", "审定数", "余额", "总额", "总计"]
                    ):
                        # 同行或下一行找数值
                        for j in range(i+1, len(row)):
                            if isinstance(row[j], (int, float)):
                                figures[f"{cell}({sn})"] = float(row[j])
                                break
        return figures
    except Exception: return {}

