#!/usr/bin/env python3
"""RAG 知识检索引擎 - TF-IDF 关键词检索（免 GPU，CPU 友好）"""
from __future__ import annotations
import json, logging, os, pickle, re, warnings
from pathlib import Path
from typing import Dict, List, Optional

# 抑制 pdfminer.six 的 FontBBox 等字体解析警告（这些 PDF 字体描述符损坏不影响业务逻辑）
warnings.filterwarnings("ignore", category=UserWarning, module="pdfminer")
logging.getLogger("pdfminer").setLevel(logging.ERROR)
logging.getLogger("pdfplumber").setLevel(logging.ERROR)

KNOWLEDGE_BASE_DIR = Path("D:/审计准则与法规文件整理")

# 🔒 这些目录包含模板文件，永不索引（防止格式泄露 + 数据安全）
RAG_SKIP_PATHS = [
    "C_底稿模板",
    "06_报告与格式规范",
    "询证函范本",
]
CACHE_DIR = Path(__file__).resolve().parent.parent / "data" / "rag_cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
CHUNK_SIZE = 1500; TOP_K = 15

_vectorizer = None; _doc_matrix = None; _doc_chunks: List[Dict] = []

def _read_pdf(fp):
    """读取 PDF 文本，静默处理字体描述符损坏等常见 PDF 缺陷"""
    try:
        import pdfplumber
        # 使用 laparams 提升对损坏 PDF 的兼容性
        with pdfplumber.open(fp, laparams={"detect_vertical": False}) as pdf:
            texts = []
            for p in pdf.pages:
                try:
                    t = p.extract_text()
                    if t:
                        texts.append(t)
                except Exception:
                    # 单页提取失败不影响其他页
                    continue
            return "\n".join(texts)
    except Exception:
        return ""

def _read_docx(fp):
    try:
        from docx import Document
        return "\n".join(p.text for p in Document(fp).paragraphs)
    except: return ""

def _read_doc_legacy(fp):
    """读取旧版 .doc 文件"""
    # 方法1：尝试 antiword（Linux）
    try:
        import subprocess
        r = subprocess.run(["antiword", str(fp)], capture_output=True, text=True, timeout=10)
        if r.returncode == 0 and r.stdout.strip():
            return r.stdout
    except Exception:
        pass

    # 方法2：Windows COM 自动化（Word 或 WPS）
    try:
        import pythoncom
        pythoncom.CoInitialize()
        try:
            from win32com import client
            word = None
            for prog_id in ["Word.Application", "WPS.Application", "KWPS.Application", "Kingsoft.WPS"]:
                try:
                    word = client.Dispatch(prog_id)
                    break
                except Exception:
                    continue
            if word is None:
                raise RuntimeError("无可用办公软件")
            word.Visible = False
            word.DisplayAlerts = 0
            try:
                doc = word.Documents.Open(str(Path(fp).resolve()))
                text = doc.Content.Text
                doc.Close()
                return text
            finally:
                word.Quit()
        finally:
            pythoncom.CoUninitialize()
    except Exception:
        pass

    # 方法3：当做纯文本硬读，过滤二进制乱码
    try:
        raw = open(fp, "rb").read()
        # 尝试 utf-16（Word 内部常用编码）
        for enc in ["utf-16-le", "utf-8", "gbk", "gb2312", "latin-1"]:
            try:
                text = raw.decode(enc, errors="ignore")
                # 过滤掉大量不可打印字符的垃圾段
                clean = "".join(
                    c for c in text
                    if c.isprintable() or c in "\n\r\t"
                )
                if len(clean) > 100:
                    return clean
            except Exception:
                continue
    except Exception:
        pass

    return ""

def _read_xlsx(fp):
    """读取 Excel 全部单元格文本（.xlsx）"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        texts = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            texts.append(f"[Sheet: {sn}]")
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join(str(v) for v in row if v is not None)
                if row_text.strip():
                    texts.append(row_text)
        return "\n".join(texts)
    except: return ""

def _read_xls(fp):
    """读取旧版 .xls 文件"""
    try:
        import pandas as pd
        xls = pd.ExcelFile(fp); texts = []
        for sn in xls.sheet_names:
            df = pd.read_excel(fp, sheet_name=sn)
            texts.append(f"[Sheet: {sn}]")
            for _, row in df.iterrows():
                row_text = " ".join(str(v) for v in row if pd.notna(v))
                if row_text.strip():
                    texts.append(row_text)
        return "\n".join(texts)
    except: return ""

def _read_txt(fp):
    for enc in ["utf-8","gbk"]:
        try: return open(fp,encoding=enc).read()
        except: pass
    return ""

def _read_file(fp):
    ext = Path(fp).suffix.lower()
    if ext == ".pdf": return _read_pdf(fp)
    elif ext == ".docx": return _read_docx(fp)
    elif ext == ".doc": return _read_doc_legacy(fp) or _read_docx(fp) or _read_txt(fp)
    elif ext == ".xlsx": return _read_xlsx(fp)
    elif ext == ".xls": return _read_xls(fp)
    elif ext in (".txt", ".md"): return _read_txt(fp)
    return ""

# 知识来源分层权重
SOURCE_WEIGHTS = {
    "01_中注协审计准则体系": 1.5,
    "02_财政部企业会计准则体系": 1.5,
    "03_事务所内部文件": 1.2,
    "04_法律法规": 1.3,
    "05_行业专项政策": 1.0,
    "06_报告与格式规范": 0.9,
}

def _get_source_category(src: str) -> str:
    for cat in SOURCE_WEIGHTS:
        if src.startswith(cat) or cat in src: return cat
    return "其他"

def _chunk_text(text, source):
    """智能分块：按段落边界，保留语义完整性"""
    chunks = []
    paras = [p.strip() for p in text.split("\n\n") if len(p.strip()) > 10]
    cur = ""
    for p in paras:
        combined = cur + "\n" + p if cur else p
        if len(combined) > CHUNK_SIZE and cur:
            chunks.append({
                "text": cur.strip(), "source": source,
                "category": _get_source_category(source)
            })
            cur = p if len(p) <= CHUNK_SIZE else p[:CHUNK_SIZE]
        else:
            cur = combined
    if cur.strip():
        chunks.append({
            "text": cur.strip(), "source": source,
            "category": _get_source_category(source)
        })
    return chunks


def _tokenize(text):
    try:
        import jieba; return " ".join(jieba.cut(text))
    except ImportError: return " ".join(text)

def build_index(force=False):
    global _vectorizer, _doc_matrix, _doc_chunks
    cache = CACHE_DIR / "tfidf_index.pkl"
    if not force and cache.exists():
        try:
            with open(cache,"rb") as f:
                d = pickle.load(f)
                _vectorizer, _doc_matrix, _doc_chunks = d["v"], d["m"], d["c"]
                print(f"[RAG] 缓存加载: {len(_doc_chunks)} 块")
                return len(_doc_chunks)
        except: pass
    from sklearn.feature_extraction.text import TfidfVectorizer
    chunks = load_all_documents()
    if not chunks: return 0
    texts = [_tokenize(c["text"]) for c in chunks]
    # 过滤掉分词后完全为空的文本块（如损坏 PDF 提取出的空白内容）
    non_empty = [(t, c) for t, c in zip(texts, chunks) if t.strip()]
    if not non_empty:
        print("[RAG] 警告: 所有文档分词后均为空，跳过索引构建（知识库 PDF 可能存在字体损坏）")
        _vectorizer = None
        _doc_matrix = None
        _doc_chunks = []
        return 0
    texts, chunks = zip(*non_empty)
    texts, chunks = list(texts), list(chunks)
    try:
        _vectorizer = TfidfVectorizer(max_features=10000)
        _doc_matrix = _vectorizer.fit_transform(texts)
        _doc_chunks = chunks
        with open(cache,"wb") as f:
            pickle.dump({"v":_vectorizer,"m":_doc_matrix,"c":_doc_chunks}, f)
        print(f"[RAG] 索引完成: {len(_doc_chunks)} 块")
        return len(_doc_chunks)
    except ValueError as e:
        # 捕获 "empty vocabulary" 等 sklearn 错误
        print(f"[RAG] 索引构建失败（向量化错误）: {e}，RAG 将降级为空结果")
        _vectorizer = None
        _doc_matrix = None
        _doc_chunks = []
        return 0

def retrieve(query, top_k=TOP_K):
    """检索相关法规片段。如果索引未就绪，返回空（不尝试构建，避免阻塞事件循环）。"""
    if _vectorizer is None or _doc_matrix is None or _doc_matrix.shape[0] == 0:
        # 尝试从缓存加载（build_index 已在启动时通过线程池预建）
        if CACHE_DIR.joinpath("tfidf_index.pkl").exists():
            build_index()
        else:
            return []
    if _vectorizer is None or _doc_matrix is None or _doc_matrix.shape[0] == 0:
        return []
    from sklearn.metrics.pairwise import cosine_similarity
    qv = _vectorizer.transform([_tokenize(query)])
    scores = cosine_similarity(qv, _doc_matrix).flatten()
    top = scores.argsort()[-top_k:][::-1]
    return [{"score":float(scores[i]),"text":_doc_chunks[i]["text"],
             "source":_doc_chunks[i]["source"],
             "category": _doc_chunks[i].get("category", "其他")}
            for i in top if scores[i]>0.01]


def hybrid_retrieve(query: str, top_k: int = TOP_K) -> List[dict]:
    """
    混合检索：向量语义 + TF-IDF关键词 双路召回，RRF 融合。
    
    优势互补：
    - 向量路：语义理解（"医保回款"→"货币资金审计"）
    - TF-IDF路：精确关键词（"函证"→命中函证准则原文）
    """
    results_a = []
    results_b = []

    # 路A：向量语义检索
    try:
        from core.rag_vector import semantic_retrieve
        results_a = semantic_retrieve(query, top_k=min(top_k * 2, 50))
    except Exception:
        pass

    # 路B：TF-IDF 关键词检索
    results_b = retrieve(query, top_k=min(top_k * 2, 50))

    # 如果只有一条路可用，直接返回
    if not results_a:
        return results_b[:top_k]
    if not results_b:
        return results_a[:top_k]

    # RRF 融合（Reciprocal Rank Fusion, k=60）
    K = 60
    merged = {}

    for rank, r in enumerate(results_a):
        key = r["text"][:200]  # 用文本前200字符去重
        merged[key] = {
            "score": 1.0 / (K + rank + 1),
            "text": r["text"],
            "source": r["source"],
            "category": r.get("category", "其他"),
            "rank_vector": rank + 1,
            "rank_tfidf": None,
        }

    for rank, r in enumerate(results_b):
        key = r["text"][:200]
        if key in merged:
            merged[key]["score"] += 1.0 / (K + rank + 1)
            merged[key]["rank_tfidf"] = rank + 1
        else:
            merged[key] = {
                "score": 1.0 / (K + rank + 1),
                "text": r["text"],
                "source": r["source"],
                "category": r.get("category", "其他"),
                "rank_vector": None,
                "rank_tfidf": rank + 1,
            }

    # 按融合分数排序，归一化到 0~1
    ranked = sorted(merged.values(), key=lambda x: x["score"], reverse=True)
    if ranked:
        max_s = ranked[0]["score"]
        for r in ranked:
            r["score"] = round(r["score"] / max_s, 4) if max_s > 0 else 0

    ranked = ranked[:top_k]

    # Reranker 精排（v3.2）：embedding 余弦相似度重排序
    ranked = _rerank(query, ranked, top_k)

    return ranked


def _rerank(query: str, candidates: List[dict], top_k: int) -> List[dict]:
    """对候选结果做 embedding 相似度重排（Qwen3-0.6B > MiniLM 降级）。

    用 rag_vector 的同一个模型对 query 和每个候选 text 计算余弦相似度，
    与 RRF 分数加权融合（RRF:sim = 0.4:0.6），重排后返回 top_k。
    """
    if not candidates or len(candidates) <= 1:
        return candidates
    try:
        from core.rag_vector import _init_model as _get_vmodel, _model as _vmodel
        _get_vmodel()
        model = _vmodel
        if model is None:
            return candidates

        from sentence_transformers.util import cos_sim
        q_emb = model.encode([query], normalize_embeddings=True, show_progress_bar=False)
        texts = [c["text"][:500] for c in candidates]
        t_embs = model.encode(texts, normalize_embeddings=True, show_progress_bar=False)
        sims = cos_sim(q_emb, t_embs)[0]

        # RRF:sim = 0.4:0.6 加权融合
        for i, c in enumerate(candidates):
            rrf_score = c.get("score", 0)
            sim_score = max(0, float(sims[i]))
            c["score"] = round(0.4 * rrf_score + 0.6 * sim_score, 4)
            c["rerank_sim"] = round(sim_score, 4)

        candidates.sort(key=lambda x: x["score"], reverse=True)
        return candidates[:top_k]
    except Exception:
        return candidates

# ═══════════════════════════════════════════════════════════════
# 法规交叉引用 1-hop 扩展（GraphRAG 轻量版：规则抽取引用边，零 LLM 成本）
# ═══════════════════════════════════════════════════════════════

_CITATION_RE = re.compile(r"《([^《》\n]{2,40}?)》")


def _extract_citations(text: str) -> List[str]:
    """从法规文本中抽取《…》书名号引用（如《中国注册会计师审计准则第1312号》）"""
    return [m.group(1).strip() for m in _CITATION_RE.finditer(text or "")]


def expand_with_citations(results: List[dict], max_extra: int = 3,
                          scan_top: int = 5) -> List[dict]:
    """对检索结果做 1-hop 引用扩展。

    命中片段正文出现《X》，且知识库中存在文件名/路径含 X 的文档时，
    追加该文档的首个片段（标记 via='citation'、score=0.30），
    解决"收入准则→应用指南→案例"式多跳引用链传统 RAG 检索不到的问题。
    索引未就绪或无引用时原样返回，绝不影响主链路。
    """
    if not results or not _doc_chunks:
        return results
    try:
        existing_keys = {r.get("text", "")[:200] for r in results}
        existing_sources = {str(r.get("source", "")) for r in results}
        cited_titles: List[str] = []
        for r in results[:scan_top]:
            for t in _extract_citations(r.get("text", "")):
                if t not in cited_titles:
                    cited_titles.append(t)
        extras = []
        for title in cited_titles:
            if len(extras) >= max_extra:
                break
            # 去掉常见前缀噪音后按"包含"匹配源文档名
            core = title.replace("中国注册会计师", "").replace("中华人民共和国", "").strip()
            if len(core) < 2:
                continue
            probe = core if len(core) <= 12 else core[:12]
            for c in _doc_chunks:
                src = str(c.get("source", ""))
                if core in src or probe in src:
                    if src in existing_sources:
                        break  # 该被引文档已在结果中，无需扩展
                    key = c["text"][:200]
                    if key in existing_keys:
                        break
                    extras.append({
                        "score": 0.30, "text": c["text"], "source": src,
                        "category": c.get("category", "其他"),
                        "via": "citation", "cited_as": f"《{title}》",
                    })
                    existing_keys.add(key)
                    existing_sources.add(src)
                    break
        if extras:
            print(f"[RAG] 引用扩展命中 {len(extras)} 条: "
                  + ", ".join(e["cited_as"] for e in extras))
        return results + extras
    except Exception as e:
        print(f"[RAG] 引用扩展异常（忽略）: {e}")
        return results


def inject_compliance_context(user_intent):
    """注入合规上下文：混合检索 > 向量 > TF-IDF，不可用时静默返回空（不阻断主流程）"""
    try:
        # 第1选择：混合检索（向量 + TF-IDF 双路融合）
        try:
            results = hybrid_retrieve(user_intent)
            engine_label = "混合检索（向量+关键词双路融合）"
        except Exception:
            # 第2选择：纯向量
            try:
                from core.rag_vector import semantic_retrieve
                results = semantic_retrieve(user_intent)
                engine_label = "向量语义检索"
            except Exception:
                # 第3选择：纯 TF-IDF
                results = retrieve(user_intent)
                engine_label = "TF-IDF关键词检索"

        if not results:
            return ""

        lines = [f"## 合规红线（{engine_label}）"]
        lines.append(f"（检索到 {len(results)} 条最相关法规/准则片段）\n")
        for i, r in enumerate(results):
            cat = r.get("category", "")
            cat_tag = f" [{cat}]" if cat else ""
            lines.append(f"### {i+1}. {r['source']}{cat_tag} (相关度:{r['score']:.4f})")
            lines.append(r["text"])
            lines.append("")
        return "\n".join(lines)
    except Exception as e:
        print(f"[RAG] inject_compliance_context 降级（非致命）: {e}")
        return ""


# ═══════════════════════════════════════════════════════════════
# 分层注入：按 DAG 算子类型加载不同知识
# ═══════════════════════════════════════════════════════════════

# 算子 → 知识分类映射
OPERATOR_KNOWLEDGE_MAP = {
    "Load":          ["03_事务所内部文件", "05_行业专项政策"],
    "RegexFilter":   ["05_行业专项政策", "03_事务所内部文件"],
    "ColumnFilter":  ["03_事务所内部文件"],
    "GroupBy":       ["03_事务所内部文件"],
    "Merge":         ["01_中注协审计准则体系", "03_事务所内部文件"],
    "Diff":          ["01_中注协审计准则体系", "03_事务所内部文件"],
    "ConditionCheck":["01_中注协审计准则体系", "04_法律法规"],
    "Export":        ["06_报告与格式规范"],
    "Aggregate":     ["03_事务所内部文件"],
    "Sort":          [],
    "Transform":     ["03_事务所内部文件"],
    "NoiseFilter":   ["03_事务所内部文件", "05_行业专项政策"],
    "Extract":       ["03_事务所内部文件"],
    "Reconcile":     ["01_中注协审计准则体系", "02_财政部企业会计准则体系"],
    "AuditAdjustment":["01_中注协审计准则体系", "02_财政部企业会计准则体系"],
}

def inject_layered_context(user_intent: str, operator_type: str = None) -> str:
    """
    分层注入：根据 DAG 算子类型，加载该算子最相关的知识子集。
    
    Args:
        user_intent: 审计师的大白话意图
        operator_type: DAG 算子名称（如 "Merge"、"RegexFilter"），为 None 时等同于通用注入

    Returns:
        合规上下文文本，可直接前置到 Prompt
    """
    try:
        target_cats = None
        if operator_type and operator_type in OPERATOR_KNOWLEDGE_MAP:
            target_cats = OPERATOR_KNOWLEDGE_MAP[operator_type]

        all_results = hybrid_retrieve(user_intent, top_k=50)
        if not all_results:
            return ""

        if target_cats:
            matched = [r for r in all_results
                       if any(bc in r.get("category", "") for bc in target_cats)]
            rest = [r for r in all_results
                    if not any(bc in r.get("category", "") for bc in target_cats)]
            core = matched[:10]
            supplement = rest[:max(0, 15 - len(core))]
            results = core + supplement
            engine_label = f"分层 [{operator_type}] -> " + ",".join(target_cats[:2])
        else:
            results = all_results[:15]
            engine_label = "通用注入"

        if not results:
            return ""

        lines = [f"## 合规红线（{engine_label}）"]
        lines.append(f"（检索到 {len(results)} 条最相关法规/准则片段）\n")
        for i, r in enumerate(results):
            cat = r.get("category", "")
            cat_tag = f" [{cat}]" if cat else ""
            lines.append(f"### {i+1}. {r['source']}{cat_tag} (相关度:{r['score']:.4f})")
            lines.append(r["text"])
            lines.append("")
        return "\n".join(lines)
    except Exception as e:
        print(f"[RAG] inject_layered_context 降级（非致命）: {e}")
        return ""

def get_status():
    return {"chunks":len(_doc_chunks),"cache":(CACHE_DIR/"tfidf_index.pkl").exists(),
            "dirs":{d.name:sum(1 for _ in d.rglob("*") if _.is_file())
                    for d in KNOWLEDGE_BASE_DIR.iterdir() if d.is_dir()} if KNOWLEDGE_BASE_DIR.exists() else {}}


def get_light_status():
    """轻量状态探针（毫秒级）：不扫描知识库目录。

    get_status() 会对知识库全盘 rglob 统计文件数，大目录下耗时可达秒级；
    前端「知识库就绪」徽标与 TopBar 健康检查高频轮询时应使用本函数。
    """
    cache_exists = (CACHE_DIR / "tfidf_index.pkl").exists()
    return {"chunks": len(_doc_chunks), "cache": cache_exists,
            "ready": bool(_doc_chunks) or cache_exists}


def load_all_documents():
    chunks = []
    if not KNOWLEDGE_BASE_DIR.exists(): return chunks
    for root, dirs, files in os.walk(str(KNOWLEDGE_BASE_DIR)):
        rel_root = os.path.relpath(root, str(KNOWLEDGE_BASE_DIR))

        # 跳过模板目录：这些是格式骨架，不是知识文本
        if any(skip in rel_root for skip in RAG_SKIP_PATHS):
            continue

        for f in files:
            fp = os.path.join(root, f)
            rel = os.path.relpath(fp, str(KNOWLEDGE_BASE_DIR))

            # .md 文件：按标题层级结构化分块
            if f.endswith(".md"):
                try:
                    from core.md_engine import chunk_by_headings
                    chunks.extend(chunk_by_headings(fp))
                    continue
                except Exception:
                    pass

            # 其他文件：按段落智能分块
            text = _read_file(fp)
            if text:
                chunks.extend(_chunk_text(text, rel))
    print(f"[RAG] 已加载 {len(chunks)} 个文本块")
    return chunks
