#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
多文件风险筛查引擎 (risk_scanner.py)
扔一堆文件 → 自动识别 → 配对/降级 → 双边/单侧异常交易特征 → 重点核查事项归并。
三条纪律：①聚合输出 ②降级优雅 ③单侧打折标注
"""

from __future__ import annotations
import os, re
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import pandas as pd


def scan_files(file_paths: List[str], out_dir: str = None) -> Dict[str, Any]:
    files = [Path(f) for f in file_paths]
    if not files:
        return {"task_cards": [], "red_flags": [], "degrade_notes": [], "stats": {}}

    # 阶段一：识别 + 配对
    file_info = _classify_files(files)
    pairs, singles, unknowns = _pair_files(file_info)
    degrade_notes = []
    for u in unknowns:
        degrade_notes.append(f"{u['name']}: 无法识别类型，仅做基础体检")
    for s in singles:
        degrade_notes.append(f"{s['name']}: 单文件({s['type']})，仅做单侧异常特征初筛")

    results = {"task_cards": [], "red_flags": [], "degrade_notes": degrade_notes,
               "stats": {"total_files": len(files), "paired": len(pairs) * 2,
                         "single": len(singles), "unknown": len(unknowns)}}

    # 阶段二：分级扫描
    all_flags = []
    for book_info, bank_info in pairs:
        try:
            from core.bank_reconcile_engine import (
                detect_book_type, auto_map_columns, normalize_to_std,
                run_bank_reconciliation, detect_red_flags
            )
            rec_result = run_bank_reconciliation(book_info["df"], bank_info["df"],
                config={"book_type": book_info["type"], "bank_type": bank_info["type"]})
            for side_df, label in [(rec_result["book_std"], f"账:{book_info['name']}"),
                                     (rec_result["bank_std"], f"银:{bank_info['name']}")]:
                flags = detect_red_flags(side_df, side=label)
                for f in flags:
                    f["proof_level"] = "对账交叉验证"
                    f["source_file"] = book_info["name"] + "+" + bank_info["name"]
                all_flags.extend(flags)
            results["stats"]["matched_pairs"] = results["stats"].get("matched_pairs", 0) + 1
        except Exception as e:
            degrade_notes.append(f"对账失败({book_info['name']}+{bank_info['name']}): {e}")

    for info in singles:
        try:
            from core.bank_reconcile_engine import (
                auto_map_columns, normalize_to_std, detect_red_flags
            )
            mapping = auto_map_columns(info["df"], info["type"])
            std_df = normalize_to_std(info["df"], mapping, info["type"], info["name"])
            flags = detect_red_flags(std_df, side=f"{info['name']}(单侧)")
            for f in flags:
                f["proof_level"] = "单侧初筛（未经对账交叉验证）"
                f["source_file"] = info["name"]
            all_flags.extend(flags)
        except Exception as e:
            degrade_notes.append(f"单侧扫描失败({info['name']}): {e}")

    for info in unknowns:
        if info["df"] is not None:
            results["stats"]["unknown_rows"] = results["stats"].get("unknown_rows", 0) + len(info["df"])

    # 阶段三：任务卡归并
    task_cards = _build_task_cards(all_flags)
    results["task_cards"] = task_cards
    results["red_flags"] = all_flags
    results["stats"]["total_flags"] = len(all_flags)
    results["stats"]["task_card_count"] = len(task_cards)

    # 阶段四：导出
    if out_dir:
        out_path = Path(out_dir) / "多文件风险筛查结果_重点核查事项.xlsx"
        _export_scan_report(out_path, results)
        results["output_file"] = str(out_path)


def _classify_files(files: List[Path]) -> List[Dict]:
    out = []
    for fp in files:
        try:
            ext = fp.suffix.lower()
            if ext in (".xlsx", ".xls"):
                df = pd.read_excel(fp, nrows=1000)
            elif ext == ".csv":
                df = pd.read_csv(fp, nrows=1000, encoding="utf-8-sig")
            else:
                out.append({"name": fp.name, "type": "unsupported", "df": None})
                continue
            from core.bank_reconcile_engine import detect_book_type
            ftype = detect_book_type(df, fp.name)
            out.append({"name": fp.name, "type": ftype, "df": df})
        except Exception as e:
            out.append({"name": fp.name, "type": "error", "df": None, "error": str(e)})
    return out


def _pair_files(file_info):
    banks = [(i, f) for i, f in enumerate(file_info) if f["type"] == "bank_statement"]
    journals = [(i, f) for i, f in enumerate(file_info) if f["type"] == "journal"]
    generics = [(i, f) for i, f in enumerate(file_info) if f["type"] == "generic_ledger"]
    unknowns = [f for f in file_info if f["type"] in ("unknown", "unsupported", "error")]
    pairs = []
    used = set()
    for bi, bank in banks:
        for ji, journal in journals:
            if ji not in used:
                pairs.append((journal, bank))
                used.add(ji); used.add(bi)
                break
    for bi, bank in banks:
        if bi not in used:
            for gi, g in generics:
                if gi not in used:
                    pairs.append((bank, g))
                    used.add(bi); used.add(gi)
                    break
    for ji, journal in journals:
        if ji not in used:
            for gi, g in generics:
                if gi not in used:
                    pairs.append((journal, g))
                    used.add(ji); used.add(gi)
                    break
    singles = [f for i, f in enumerate(file_info)
               if i not in used and f["type"] not in ("unknown", "unsupported", "error")]
    return pairs, singles, unknowns


def _build_task_cards(flags):
    cp_map = defaultdict(lambda: defaultdict(list))
    for f in flags:
        rows = f.get("rows", [])
        if not rows: continue
        cp = _extract_counterpart(f)
        ft = f.get("type", "其他")
        cp_map[cp][ft].append(f)
    cards = []
    for cp, ft_map in cp_map.items():
        for ft, fs in ft_map.items():
            total_amt = sum(abs(float(x.get("amount", 0))) for x in fs)
            proof = fs[0].get("proof_level", "") if fs else ""
            src = fs[0].get("source_file", "") if fs else ""
            cards.append({
                "对手方": cp, "旗型": ft, "笔数": len(fs),
                "涉及金额": total_amt, "证明力": proof,
                "来源文件": src, "红旗明细": fs,
            })
    cards.sort(key=lambda c: c["涉及金额"], reverse=True)
    return cards


def _extract_counterpart(flag):
    for key in ("counterpart", "对手方", "对方", "对方客户名称"):
        v = flag.get(key, "")
        if v and str(v).strip() and len(str(v).strip()) >= 2:
            return str(v).strip()[:30]
    detail = str(flag.get("detail", ""))
    m = re.search(r"对手方\[(.+?)\]", detail)
    if m: return m.group(1)[:30]
    return "(未识别对手方)"


def _export_scan_report(path: Path, results):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active; ws.title = "重点核查事项"
    hdrs = ["对手方", "特征类型", "笔数", "涉及金额", "证据充分性", "来源文件"]
    for i, h in enumerate(hdrs, 1):
        ws.cell(row=1, column=i, value=h).font = Font(bold=True)
    for ri, card in enumerate(results.get("task_cards", []), 2):
        ws.cell(row=ri, column=1, value=card["对手方"])
        ws.cell(row=ri, column=2, value=card["旗型"])
        ws.cell(row=ri, column=3, value=card["笔数"])
        ws.cell(row=ri, column=4, value=card["涉及金额"])
        pc = ws.cell(row=ri, column=5, value=card["证明力"])
        if "单侧" in card["证明力"]:
            pc.font = Font(color="FF8C00", italic=True)
        ws.cell(row=ri, column=6, value=card["来源文件"])
    for ci, w in {1: 26, 2: 18, 3: 8, 4: 16, 5: 30, 6: 30}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws2 = wb.create_sheet("扫描范围与局限")
    ws2.cell(row=1, column=1, value="扫描范围与局限").font = Font(bold=True, size=14)
    st = results.get("stats", {})
    items = [
        f"总文件: {st.get('total_files', 0)}",
        f"对账配对: {st.get('matched_pairs', 0)} 对",
        f"单侧扫描: {st.get('single', 0)} 个",
        f"无法识别: {st.get('unknown', 0)} 个",
        f"重点核查事项: {st.get('task_card_count', 0)} 项",
        f"异常特征总数: {st.get('total_flags', 0)}",
    ]
    for i, item in enumerate(items, 3):
        ws2.cell(row=i, column=1, value=item)
    ws2.cell(row=10, column=1, value="明细:").font = Font(bold=True)
    for i, note in enumerate(results.get("degrade_notes", []), 11):
        ws2.cell(row=i, column=1, value=note)
    ws2.column_dimensions["A"].width = 80

    ws3 = wb.create_sheet("异常交易特征明细")
    for i, h in enumerate(["类型", "金额", "证据充分性", "来源", "详情"], 1):
        ws3.cell(row=1, column=i, value=h).font = Font(bold=True)
    for ri, f in enumerate(results.get("red_flags", []), 2):
        ws3.cell(row=ri, column=1, value=f.get("type", ""))
        ws3.cell(row=ri, column=2, value=f.get("amount", 0))
        ws3.cell(row=ri, column=3, value=f.get("proof_level", ""))
        ws3.cell(row=ri, column=4, value=f.get("source_file", ""))
        ws3.cell(row=ri, column=5, value=str(f.get("detail", ""))[:120])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    print(f"[多文件风险筛查] 已导出: {path}")



    return results