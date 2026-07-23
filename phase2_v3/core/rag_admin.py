#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RAG 知识库后台管理 —— 增量索引重建、状态监控
当审计师下载新法规放入知识库目录后，调用 rebuild_index 即可生效，
无需重启服务。
"""
from __future__ import annotations
import os, pickle, logging
from pathlib import Path
from typing import Dict, List

logger = logging.getLogger(__name__)

KNOWLEDGE_BASE_DIR = Path("D:/审计准则与法规文件整理")
CACHE_DIR = Path(__file__).resolve().parent.parent / "data" / "rag_cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

CHUNK_SIZE = 800


def _read_pdf(fp):
    try:
        import pdfplumber
        with pdfplumber.open(fp, laparams={"detect_vertical": False}) as pdf:
            texts = []
            for p in pdf.pages:
                try:
                    t = p.extract_text()
                    if t: texts.append(t)
                except Exception: continue
            return "\n".join(texts)
    except Exception: return ""


def _read_docx(fp):
    try:
        from docx import Document
        return "\n".join(p.text for p in Document(fp).paragraphs)
    except Exception: return ""


def _read_txt(fp):
    for enc in ["utf-8", "gbk"]:
        try: return open(fp, encoding=enc).read()
        except Exception: pass
    return ""


def _read_xlsx(fp):
    """读取 Excel 全部 Sheet 的单元格文本"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        texts = []
        for sn in wb.sheetnames:
            ws = wb[sn]
            texts.append(f"[Sheet: {sn}]")
            for row in ws.iter_rows(values_only=True):
                row_text = " ".join(str(v) for v in row if v is not None)
                if row_text.strip():
                    texts.append(row_text)
        return "\n".join(texts)
    except Exception:
        return ""


def _read_xls(fp):
    """读取旧版 .xls 文件"""
    try:
        import pandas as pd
        xls = pd.ExcelFile(fp)
        texts = []
        for sn in xls.sheet_names:
            df = pd.read_excel(fp, sheet_name=sn)
            texts.append(f"[Sheet: {sn}]")
            for _, row in df.iterrows():
                row_text = " ".join(str(v) for v in row if pd.notna(v))
                if row_text.strip():
                    texts.append(row_text)
        return "\n".join(texts)
    except Exception:
        return ""


def _read_file(fp):
    ext = Path(fp).suffix.lower()
    if ext == ".pdf": return _read_pdf(fp)
    if ext in (".docx", ".doc"): return _read_docx(fp) or _read_txt(fp)
    if ext in (".xlsx",): return _read_xlsx(fp)
    if ext in (".xls",): return _read_xls(fp)
    if ext in (".txt", ".md"): return _read_txt(fp)
    return ""


def _chunk_text(text: str, source: str) -> list:
    chunks = []
    paras = [p.strip() for p in text.split("\n") if len(p.strip()) > 20]
    cur = ""
    for p in paras:
        if len(cur) + len(p) > CHUNK_SIZE and cur:
            chunks.append({"text": cur.strip(), "source": source})
            cur = p
        else:
            cur += "\n" + p if cur else p
    if cur.strip():
        chunks.append({"text": cur.strip(), "source": source})
    return chunks


def scan_directory() -> dict:
    """扫描知识库目录，返回各子目录文件统计"""
    result = {}
    if not KNOWLEDGE_BASE_DIR.exists():
        return {"error": f"知识库目录不存在: {KNOWLEDGE_BASE_DIR}"}

    for d in sorted(KNOWLEDGE_BASE_DIR.iterdir()):
        if d.is_dir():
            files = []
            for root, dirs, fnames in os.walk(str(d)):
                for f in fnames:
                    fp = Path(root) / f
                    files.append({
                        "name": f,
                        "path": str(fp.relative_to(KNOWLEDGE_BASE_DIR)),
                        "size": fp.stat().st_size,
                        "modified": fp.stat().st_mtime,
                    })
            result[d.name] = {
                "file_count": len(files),
                "total_size_mb": round(sum(f["size"] for f in files) / 1024 / 1024, 2),
                "files": files,
            }
    return result


def check_index_freshness() -> dict:
    """检查索引是否需要重建（知识库文件比缓存新）"""
    tfidf_cache = CACHE_DIR / "tfidf_index.pkl"
    vector_cache = CACHE_DIR / "vector_faiss.pkl"
    caches = {}
    
    for label, cache_file in [("tfidf", tfidf_cache), ("vector", vector_cache)]:
        if not cache_file.exists():
            caches[label] = {"status": "no_cache", "message": "索引缓存不存在，需要重建"}
            continue
        
        cache_time = cache_file.stat().st_mtime
        stale_files = []
        total_files = 0
        
        if KNOWLEDGE_BASE_DIR.exists():
            for root, dirs, fnames in os.walk(str(KNOWLEDGE_BASE_DIR)):
                for f in fnames:
                    total_files += 1
                    fp = Path(root) / f
                    if fp.stat().st_mtime > cache_time:
                        stale_files.append(str(fp.relative_to(KNOWLEDGE_BASE_DIR)))
        
        caches[label] = {
            "status": "stale" if stale_files else "fresh",
            "cache_time": cache_time,
            "total_files": total_files,
            "stale_count": len(stale_files),
            "stale_files": stale_files[:10],
        }
    
    return caches


def rebuild_index(force: bool = False) -> dict:
    """完全重建 RAG 索引，返回统计信息（同时重建 TF-IDF 和向量索引）"""
    from core.rag_engine import build_index, get_status

    if not force:
        freshness = check_index_freshness()
        # freshness 现在是 {"tfidf": {...}, "vector": {...}} 格式
        tfidf_stale = freshness.get("tfidf", {}).get("status") != "fresh"
        vector_stale = freshness.get("vector", {}).get("status") != "fresh"
        if not tfidf_stale and not vector_stale:
            return {"status": "skipped", "message": "所有索引已是最新", "stats": get_status()}

    # 强制重建 TF-IDF
    try:
        chunk_count = build_index(force=True)
    except Exception as e:
        return {"status": "error", "message": f"TF-IDF重建失败: {e}"}

    # 同时重建向量索引
    vector_status = "skipped"
    try:
        from core.rag_vector import build_vector_index
        vcount = build_vector_index(force=True)
        vector_status = f"OK ({vcount} blocks)"
    except Exception as e:
        vector_status = f"failed: {e}"

    status = get_status()
    return {
        "status": "success",
        "chunk_count": chunk_count,
        "vector_index": vector_status,
        "stats": status,
    }


def get_rag_status() -> dict:
    """获取 RAG 系统完整状态"""
    from core.rag_engine import get_status as _get_status
    status = _get_status()
    directory_scan = scan_directory()
    freshness = check_index_freshness()
    return {
        "index": status,
        "directories": directory_scan,
        "freshness": freshness,
    }


if __name__ == "__main__":
    # CLI 入口：python -m core.rag_admin [scan|rebuild|status|freshness]
    import sys
    cmd = sys.argv[1] if len(sys.argv) > 1 else "status"
    import json

    if cmd == "scan":
        print(json.dumps(scan_directory(), ensure_ascii=False, indent=2))
    elif cmd == "rebuild":
        print(json.dumps(rebuild_index(force=True), ensure_ascii=False, indent=2))
    elif cmd == "freshness":
        print(json.dumps(check_index_freshness(), ensure_ascii=False, indent=2))
    else:
        print(json.dumps(get_rag_status(), ensure_ascii=False, indent=2))
