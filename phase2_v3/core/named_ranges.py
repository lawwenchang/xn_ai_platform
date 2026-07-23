#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
命名区域模板写入引擎 (named_ranges.py) —— 白皮书 §4.3.1 底稿引擎核心
========================================================================
审计底稿模板的"命名区域（Named Ranges）"写入规范：

    模板由质控部门预先定义命名区域（如 "audit_conclusion", "diff_total"）。
    引擎仅通过命名区域名称写入数据，杜绝硬编码单元格坐标导致的串行错位。

核心能力：
    1. scan_template()   → 扫描模板，列出所有命名区域及其引用坐标
    2. map_data()        → 将结构化数据按名称匹配到命名区域
    3. write_values()    → 写入数据（保留原始类型，不转字符串）
    4. validate_missing()→ 检查哪些命名区域未被填充（防止底稿缺项）
    5. batch_fill()      → 一行调用：模板 + 数据 → 输出底稿 + 校验通过

与 template_engine.fill_xlsx 的区别：
    fill_xlsx 通过 {{key}} 占位符匹配（适合简单填充），
    named_ranges 通过 openpyxl defined_names 匹配（适合中注协标准底稿模板），
    两者互补，可同时使用。
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook


@dataclass
class NamedRangeInfo:
    """一个命名区域的元信息"""
    name: str                        # 如 "audit_conclusion"、"diff_detail_header"
    sheet: str                       # 所在工作表名
    cell_ref: str                    # 单元格引用，如 "$A$1" 或 "$B$2:$D$20"
    is_range: bool = False           # 是否为区域（多单元格）
    row_count: int = 1               # 区域行数（单格为1）
    description: str = ""            # 从注释提取的描述（可选）

    def __repr__(self):
        return f"{self.name} → {self.sheet}!{self.cell_ref}"


def scan_template(template_path: str) -> List[NamedRangeInfo]:
    """
    扫描 .xlsx 模板，提取所有命名区域。

    Returns:
        命名区域信息列表，按 sheet 和 name 排序。
    """
    wb = load_workbook(template_path, data_only=False)
    results = []

    for name_obj in _iter_defined_names(wb):
        info = _parse_defined_name(name_obj, wb)
        if info:
            results.append(info)

    wb.close()
    results.sort(key=lambda x: (x.sheet, x.name))
    return results


def _iter_defined_names(wb) -> list:
    """兼容 openpyxl 新旧 API，遍历 workbook/worksheet 级命名区域。

    - openpyxl >= 3.1: wb.defined_names 是 DefinedNameDict（dict 子类），
      值即 DefinedName 对象；worksheet 级命名区域存于 ws.defined_names。
    - openpyxl < 3.1: wb.defined_names 是 DefinedNameList，
      通过 .definedName 属性访问。
    """
    objs = []
    dns = wb.defined_names
    if hasattr(dns, "definedName"):          # 旧版 API
        try:
            objs.extend(list(dns.definedName))
        except Exception:
            pass
    elif hasattr(dns, "values"):             # 新版 dict-like API
        try:
            objs.extend(list(dns.values()))
        except Exception:
            pass
    # worksheet 级命名区域（新版 API 才有；旧版已含在 workbook 级里）
    for ws in getattr(wb, "worksheets", []):
        ws_dns = getattr(ws, "defined_names", None)
        if ws_dns is not None and hasattr(ws_dns, "values"):
            try:
                objs.extend(list(ws_dns.values()))
            except Exception:
                pass
    # 去重（按 name），保持顺序
    seen, uniq = set(), []
    for o in objs:
        n = getattr(o, "name", None)
        if n and n not in seen:
            seen.add(n)
            uniq.append(o)
    return uniq



def _parse_defined_name(name_obj, wb) -> Optional[NamedRangeInfo]:
    """解析一个 openpyxl DefinedName 对象"""
    try:
        destinations = list(name_obj.destinations)
    except Exception:
        return None

    if not destinations:
        return None

    sheet_name, cell_ref = destinations[0]
    cells = [c.strip() for c in cell_ref.split(",")]
    is_range = ":" in cells[0] or len(cells) > 1
    row_count = 1
    if ":" in cells[0]:
        parts = cells[0].split(":")
        try:
            start = re.sub(r"\D", "", parts[0])
            end = re.sub(r"\D", "", parts[1])
            if start and end:
                row_count = int(end) - int(start) + 1
        except Exception:
            pass

    desc = ""
    if hasattr(name_obj, "comment") and name_obj.comment:
        desc = str(name_obj.comment)

    return NamedRangeInfo(
        name=name_obj.name,
        sheet=sheet_name,
        cell_ref=cell_ref,
        is_range=is_range,
        row_count=row_count,
        description=desc,
    )


def map_data(data: Dict[str, Any], named_ranges: List[NamedRangeInfo]) -> Dict[str, Tuple[NamedRangeInfo, Any]]:
    """
    将数据按名称匹配到命名区域。

    匹配规则：
    1. 精确键名匹配（data["audit_conclusion"] → named_range "audit_conclusion"）
    2. 模糊匹配（data["结论"] → named_range 包含 "conclusion" 或 "结论"）
    3. 嵌套展开（data["audit.conclusion"] → named_range "audit_conclusion"）

    Returns:
        {name: (NamedRangeInfo, value)} 映射表
    """
    import re

    # 精确匹配表
    nr_map = {nr.name: nr for nr in named_ranges}

    # 展开嵌套键
    flat_data = {}
    for k, v in data.items():
        if isinstance(v, dict):
            for sub_k, sub_v in v.items():
                flat_data[f"{k}_{sub_k}"] = sub_v
                flat_data[f"{k}.{sub_k}"] = sub_v
        else:
            flat_data[k] = v

    result = {}
    for key, value in flat_data.items():
        # 精确匹配
        if key in nr_map:
            result[key] = (nr_map[key], value)
            continue

        # 模糊匹配（key 含下划线或点号时分段匹配）
        for nr_name in nr_map:
            if key.replace(".", "_").replace(" ", "_").lower() == nr_name.lower():
                result[nr_name] = (nr_map[nr_name], value)
                break

    return result


def write_values(
    template_path: str,
    mapped: Dict[str, Tuple[NamedRangeInfo, Any]],
    output_path: str,
) -> int:
    """
    将匹配结果写入模板副本。

    - 单格命名区域 → 直接赋值（保留原始类型）
    - 多行区域 → 按行展开写入第一列

    Returns:
        成功写入的命名区域数量
    """
    wb = load_workbook(template_path)
    count = 0

    for name, (nr_info, value) in mapped.items():
        ws = wb[nr_info.sheet] if nr_info.sheet in wb.sheetnames else None
        if ws is None:
            continue

        cell_ref = nr_info.cell_ref.replace("$", "")
        if ":" in cell_ref:
            # 多单元格区域
            start, end = cell_ref.split(":")
            col_letter = re.sub(r"\d", "", start)
            start_row = int(re.sub(r"\D", "", start))
            cells_addressed = 0
            if isinstance(value, (list, tuple)):
                for i, item in enumerate(value):
                    if start_row + i > int(re.sub(r"\D", "", end)):
                        break
                    cell = ws[f"{col_letter}{start_row + i}"]
                    _set_cell_value(cell, item)
                    cells_addressed += 1
            else:
                cell = ws.cell(row=start_row, column=_col_to_num(col_letter))
                _set_cell_value(cell, value)
                cells_addressed = 1
            count += cells_addressed
        else:
            # 单格
            cell = ws[cell_ref]
            _set_cell_value(cell, value)
            count += 1

    wb.save(output_path)
    wb.close()
    return count


def _set_cell_value(cell, value):
    """设置单元格值，保留原始类型"""
    if isinstance(value, (int, float)):
        cell.value = value
    elif isinstance(value, bool):
        cell.value = value
    elif isinstance(value, str) and value.strip() == "":
        return  # 空字符串不覆盖，保留模板原文
    else:
        cell.value = str(value)


def validate_missing(mapped: Dict[str, Tuple[NamedRangeInfo, Any]],
                     all_named_ranges: List[NamedRangeInfo]) -> List[NamedRangeInfo]:
    """检查哪些命名区域未被填充，返回缺失列表"""
    written_names = set(mapped.keys())
    return [nr for nr in all_named_ranges if nr.name not in written_names]


def batch_fill(
    template_path: str,
    data: Dict[str, Any],
    output_path: str = "",
    strict: bool = False,
) -> Dict[str, Any]:
    """
    一行调用：模板 + 数据 → 输出底稿 + 校验报告。

    Args:
        template_path: .xlsx 模板路径
        data: 要填入的结构化数据
        output_path: 输出路径（默认与模板同目录加 _filled 后缀）
        strict: True 时有缺失命名区域则抛出异常

    Returns:
        {"written": int, "missing": [...], "named_ranges_total": int, "output": str}
    """
    if not output_path:
        tp = Path(template_path)
        output_path = str(tp.parent / f"{tp.stem}_filled{tp.suffix}")

    nr_list = scan_template(template_path)
    mapped = map_data(data, nr_list)
    written = write_values(template_path, mapped, output_path)
    missing = validate_missing(mapped, nr_list)

    if strict and missing:
        names = ", ".join(nr.name for nr in missing)
        raise ValueError(f"以下命名区域未填充: {names}")

    return {
        "written": written,
        "missing": [nr.name for nr in missing],
        "named_ranges_total": len(nr_list),
        "mapped_count": len(mapped),
        "output": output_path,
    }


def _col_to_num(col_letter: str) -> int:
    """列字母转数字：A→1, Z→26, AA→27"""
    n = 0
    for c in col_letter.upper():
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n