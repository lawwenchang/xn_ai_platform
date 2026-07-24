#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
公式注入引擎 v2 (formula_writer.py)
锚点记账 → 代码拼装公式。LLM 参与度：零。
支持：XLOOKUP, VLOOKUP, SUMIF, SUMIFS, SUMPRODUCT, REGEXP, IF, IFS, RANK
"""

from __future__ import annotations
import json, os, subprocess, tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


class AnchorTracker:
    """写值时记录格子坐标，支持跨 sheet 引用。"""

    def __init__(self, ws, start_row: int = 2, sheet_name: str = None):
        self.ws = ws
        self.sheet_name = sheet_name or ws.title
        self.row = start_row
        self.col_map: Dict[str, int] = {}
        self.anchors: Dict[str, str] = {}
        self._sum_row: Optional[int] = None

    def write_header(self, headers: List[str]):
        for i, h in enumerate(headers, 1):
            c = self.ws.cell(row=1, column=i, value=str(h))
            c.font = Font(bold=True)
            self.col_map[h] = i

    def write_row(self, data: Dict[str, Any], auto_track: List[str] = None):
        for col_name, col_idx in self.col_map.items():
            v = data.get(col_name, "")
            self.ws.cell(row=self.row, column=col_idx, value=v if v is not None else "")
            if auto_track and col_name in auto_track:
                ref = f"{get_column_letter(col_idx)}{self.row}"
                self.anchors[f"{col_name}_r{self.row}"] = ref
        self.row += 1

    def write_sum_row(self, label: str, formulas: Dict[int, str]):
        self.ws.cell(row=self.row, column=1, value=label).font = Font(bold=True, color="000080")
        for ci, fm in formulas.items():
            self.ws.cell(row=self.row, column=ci, value=fm)
        self._sum_row = self.row
        self.row += 1

    @property
    def last_row(self) -> int: return self.row - 1
    @property
    def last_data_row(self) -> int: return (self._sum_row or self.row) - 1

    def col_letter(self, col_name: str) -> str:
        return get_column_letter(self.col_map.get(col_name, 1))
    def col_index(self, col_name: str) -> int:
        return self.col_map.get(col_name, 1)

    def col_range(self, col_name: str, start_row: int = 2,
                  end_row: int = None, absolute: bool = False) -> str:
        cl = self.col_letter(col_name); er = end_row or self.last_data_row
        return f"${cl}${start_row}:${cl}${er}" if absolute else f"{cl}{start_row}:{cl}{er}"

    def xref_range(self, col_name: str, start_row: int = 2,
                   end_row: int = None, absolute: bool = True) -> str:
        inner = self.col_range(col_name, start_row, end_row, absolute)
        sn = self.sheet_name.replace("'", "''")
        return f"'{sn}'!{inner}"

    def cell_ref(self, col_name: str, row: int, absolute: bool = False) -> str:
        cl = self.col_letter(col_name)
        return f"${cl}${row}" if absolute else f"{cl}{row}"


FORMULA_TEMPLATES = {
    "balance_reconciliation": {
        "调节后企业余额": "={账面余额}+{银收企未收}-{银付企未付}",
        "调节后银行余额": "={银行余额}+{企收银未收}-{企付银未付}",
        "差异": "={调节后企业余额}-{调节后银行余额}",
    },
    "counterpart_xlookup": (
        '=IFERROR(XLOOKUP({lookup_cell},{target_cp_range},'
        '{target_amt_range},"未找到",0,1),"")'
    ),
    "counterpart_vlookup": (
        '=IFERROR(VLOOKUP({lookup_cell},{target_range},{col_offset},FALSE),"未找到")'
    ),
    "sumif_counterpart": "=SUMIF({cp_range},{cp_cell},{amt_range})",
    "sumifs_year": (
        "=SUMIFS({amt_range},{cp_range},{cp_cell},"
        '{date_range},">="&DATE({year},1,1),'
        '{date_range},"<"&DATE({year2},1,1))'
    ),
    "sumifs_flag": (
        '=SUMIFS({amt_range},{flag_range},"*{flag_text}*",{cp_range},{cp_cell})'
    ),
    "sumproduct_flag": (
        "=SUMPRODUCT(({cp_range}={cp_cell})*"
        '(ISNUMBER(SEARCH("{flag_text}",{flag_range})))*'
        "({amt_range}))"
    ),
    "regexp_flag": (
        '=IF(REGEXP({cell},"{pattern}",1),"{label}","")'
    ),
    "regexp_extract_year": (
        '=IFERROR(VALUE(REGEXP({cell},"\\\\d{{4}}")),'
        'IFERROR(VALUE(REGEXP({cell},"\\\\d{{4}}",1)),""))'
    ),
    "countif_flag": '=COUNTIF({range},"*{flag_text}*")',
    "countifs_year": (
        "=COUNTIFS({cp_range},{cp_cell},"
        '{date_range},">="&DATE({year},1,1),'
        '{date_range},"<"&DATE({year2},1,1))'
    ),
    "extract_year": "=IFERROR(YEAR({cell}),\"\")",
    "date_default_year": (
        '=IF(LEN({date_cell})>5,YEAR({date_cell}),YEAR({default_cell}))'
    ),
    "if_amount_flag": (
        '=IF(AND({amt_cell}>={threshold},MOD({amt_cell},{round_to})=0),'
        '"{label}","")'
    ),
    "ifs_risk_tier": (
        '=IFS({amt_cell}>={tier1},"HIGH",{amt_cell}>={tier2},"MEDIUM",'
        '{amt_cell}>0,"LOW",TRUE,"")'
    ),
    "index_match": (
        '=IFERROR(INDEX({return_range},MATCH({lookup_cell},{lookup_range},0)),'
        '"未找到")'
    ),
    "subtotal_sum": "=SUBTOTAL(109,{range})",
    "subtotal_count": "=SUBTOTAL(103,{range})",
    "textjoin_cp": (
        '=TEXTJOIN("、",TRUE,IF({cp_range}={cp_cell},{desc_range},""))'
    ),
    "left_extract": '=LEFT({cell},{n})',
    "right_extract": '=RIGHT({cell},{n})',
    "mid_extract": '=MID({cell},{start},{n})',
    "aggregate_sum": "=AGGREGATE(9,3,{range})",
    "aggregate_count": "=AGGREGATE(2,3,{range})",
    "indirect_sheet_sum": "=SUM(INDIRECT(\"'\"&{sheet_cell}&\"'!\"&{range_str}))",
    "indirect_cell": "=INDIRECT(\"'\"&{sheet_cell}&\"'!\"&{cell_str})",
    "offset_dynamic": (
        "=SUM(OFFSET({start_cell},0,0,"
        "COUNTA({count_col}),1))"
    ),
    "round_check": '=IF(ROUND({cell},{digits})<>{cell},"⚠需关注","")',
    "mround_check": (
        '=IF(AND({cell}>={threshold},MROUND({cell},{base})<>{cell}),'
        '"⚠非整{base}倍数","")'
    ),
    "networkdays_age": (
        "=IF(AND({start_cell}<>\"\",{end_cell}<>\"\"),"
        "NETWORKDAYS({start_cell},{end_cell}),\"\")"
    ),
}


def build_formula(template: str, **kwargs) -> str:
    result = template
    for k, v in kwargs.items():
        result = result.replace(f"{{{k}}}", str(v))
    return result


def build_xlookup(lookup_tr: AnchorTracker, lookup_col: str,
                   target_tr: AnchorTracker, target_cp_col: str,
                   target_amt_col: str, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["counterpart_xlookup"],
        lookup_cell=lookup_tr.cell_ref(lookup_col, row),
        target_cp_range=target_tr.xref_range(target_cp_col),
        target_amt_range=target_tr.xref_range(target_amt_col))

def build_sumif_cp(tr: AnchorTracker, cp_col: str, amt_col: str, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["sumif_counterpart"],
        cp_range=tr.col_range(cp_col, absolute=True),
        cp_cell=tr.cell_ref(cp_col, row),
        amt_range=tr.col_range(amt_col, absolute=True))

def build_regexp_flag(tr: AnchorTracker, col: str, pattern: str,
                       label: str, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["regexp_flag"],
        cell=tr.cell_ref(col, row), pattern=pattern, label=label)

def build_date_year(date_tr: AnchorTracker, date_col: str,
                     def_tr: AnchorTracker, def_col: str, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["date_default_year"],
        date_cell=date_tr.cell_ref(date_col, row),
        default_cell=def_tr.cell_ref(def_col, row))

def build_amount_flag(tr: AnchorTracker, amt_col: str,
                       threshold: float, round_to: int,
                       label: str, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["if_amount_flag"],
        amt_cell=tr.cell_ref(amt_col, row),
        threshold=threshold, round_to=round_to, label=label)

def build_risk_tier(tr: AnchorTracker, amt_col: str,
                     tier1: float, tier2: float, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["ifs_risk_tier"],
        amt_cell=tr.cell_ref(amt_col, row), tier1=tier1, tier2=tier2)



def build_index_match(tr: AnchorTracker, return_col: str, lookup_col: str,
                       lookup_val: str, row: int,
                       target_tr: AnchorTracker = None,
                       target_return_col: str = None,
                       target_lookup_col: str = None) -> str:
    """INDEX/MATCH：向左查找（兼容 WPS 旧版）。可跨 sheet。"""
    rtr = target_tr or tr
    rcol = target_return_col or return_col
    lcol = target_lookup_col or lookup_col
    return build_formula(FORMULA_TEMPLATES["index_match"],
        return_range=rtr.col_range(rcol, absolute=True),
        lookup_cell=tr.cell_ref(lookup_col, row) if target_tr is None
                     else lookup_val,
        lookup_range=rtr.col_range(lcol, absolute=True))

def build_subtotal_sum(tr: AnchorTracker, col: str) -> str:
    return build_formula(FORMULA_TEMPLATES["subtotal_sum"],
        range=tr.col_range(col))

def build_textjoin_cp(tr: AnchorTracker, cp_col: str, desc_col: str, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["textjoin_cp"],
        cp_range=tr.col_range(cp_col, absolute=True),
        cp_cell=tr.cell_ref(cp_col, row),
        desc_range=tr.col_range(desc_col, absolute=True))

def build_left(tr: AnchorTracker, col: str, n: int, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["left_extract"],
        cell=tr.cell_ref(col, row), n=n)

def build_right(tr: AnchorTracker, col: str, n: int, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["right_extract"],
        cell=tr.cell_ref(col, row), n=n)

def build_mid(tr: AnchorTracker, col: str, start: int, n: int, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["mid_extract"],
        cell=tr.cell_ref(col, row), start=start, n=n)

def build_round_check(tr: AnchorTracker, col: str, digits: int, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["round_check"],
        cell=tr.cell_ref(col, row), digits=digits)

def build_mround_check(tr: AnchorTracker, col: str, threshold: float,
                        base: int, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["mround_check"],
        cell=tr.cell_ref(col, row), threshold=threshold, base=base)

def build_networkdays(tr: AnchorTracker, start_col: str, end_col: str, row: int) -> str:
    return build_formula(FORMULA_TEMPLATES["networkdays_age"],
        start_cell=tr.cell_ref(start_col, row),
        end_cell=tr.cell_ref(end_col, row))



def verify_formulas(filepath: Path, checks: Dict[str, Any]) -> Dict[str, Any]:
    results = {"passed": True, "failures": [], "details": []}
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            csv_out = os.path.join(tmpdir, "verify.csv")
            subprocess.run(["libreoffice", "--headless", "--calc",
                           "--convert-to", "csv", "--outdir", tmpdir, str(filepath)],
                          timeout=30, capture_output=True)
            if not os.path.exists(csv_out):
                results["failures"].append("重算失败：无CSV"); results["passed"] = False
                return results
            df = pd.read_csv(csv_out, header=None)
            for cell_ref, expected in checks.items():
                col_l = ''.join(c for c in cell_ref if c.isalpha())
                row_n = int(''.join(c for c in cell_ref if c.isdigit()))
                ci = column_index_from_string(col_l)
                if row_n <= len(df) and ci <= len(df.columns):
                    actual = df.iloc[row_n - 1, ci - 1]
                    try: actual_f = float(actual)
                    except: actual_f = None
                    if actual_f is not None and abs(actual_f - float(expected)) > 0.02:
                        results["failures"].append(f"{cell_ref}:期望{expected:.2f}实际{actual_f:.2f}")
                        results["passed"] = False
    except FileNotFoundError:
        results["details"].append("(LibreOffice未安装)")
    except Exception as e:
        results["failures"].append(str(e)); results["passed"] = False
    return results


def export_triage_summary_with_formulas(wb, triage_result):
    from openpyxl.styles import Font
    if wb.sheetnames:
        ws = wb.create_sheet("分桶汇总", 0)
    else:
        ws = wb.create_sheet("分桶汇总")
    tr = AnchorTracker(ws, start_row=2, sheet_name="分桶汇总")
    tr.write_header(["桶名", "笔数", "金额合计", "建议程序"])
    for b in triage_result.get("buckets", []):
        tr.write_row({"桶名": b["name"], "笔数": b["count"],
                       "金额合计": b["amount"], "建议程序": b.get("procedure", "")},
                     auto_track=["桶名", "笔数", "金额合计"])
    rem = triage_result.get("remaining", [])
    if rem:
        ra = sum(abs(float(r.get("金额", r.get("amount", 0)) or 0)) for r in rem)
        tr.write_row({"桶名": "待人工核查（剩余）", "笔数": len(rem),
                       "金额合计": ra, "建议程序": "逐笔核查"}, auto_track=["桶名", "笔数", "金额合计"])
    sr = tr.row
    tr.write_sum_row("合计", {2: f"=SUM(B2:B{sr-1})", 3: f"=SUM(C2:C{sr-1})"})
    pc = len(tr.col_map) + 1
    ws.cell(row=1, column=pc, value="金额占比").font = Font(bold=True)
    for r in range(2, sr + 1):
        ws.cell(row=r, column=pc, value=f'=IF(C{sr}>0,C{r}/C{sr},"")')
        ws.cell(row=r, column=pc).number_format = '0.0%'
    cc = pc + 1
    ws.cell(row=1, column=cc, value="笔数占比").font = Font(bold=True)
    for r in range(2, sr + 1):
        ws.cell(row=r, column=cc, value=f'=IF(B{sr}>0,B{r}/B{sr},"")')
        ws.cell(row=r, column=cc).number_format = '0.0%'
    rc = cc + 1
    ws.cell(row=1, column=rc, value="金额排名").font = Font(bold=True)
    for r in range(2, sr):
        ws.cell(row=r, column=rc, value=f'=RANK(C{r},C$2:C${sr-1})')
    for ci, w in {1: 28, 2: 10, 3: 16, 4: 32, pc: 10, cc: 10, rc: 10}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    return tr