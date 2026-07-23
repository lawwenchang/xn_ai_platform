#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 审计工作底稿引擎
基于模板 + 命名区域写入数据 → 输出标准化底稿
"""

from pathlib import Path
from typing import Dict, List, Any, Optional

TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "data" / "templates"


def generate_workpaper(
    data: Dict[str, Any],
    output_path: str,
    template_name: str = "5.1审计报告附表Excel排版示范案例.xlsx",
) -> str:
    """
    生成 Excel 工作底稿
    data: 包含以下键的字典
      - diff_detail: 差异明细列表 [{机构, 筛选金额, 回款表金额, 差额, 差额比例}]
      - match_stats: 匹配统计
      - strategy_name: 策略名称
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "审计底稿"

    # 样式定义
    header_font = Font(name="宋体", size=11, bold=True)
    normal_font = Font(name="宋体", size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    money_fmt = '#,##0.00'

    # 标题行
    ws.merge_cells("A1:F1")
    ws["A1"] = "审计工作底稿"
    ws["A1"].font = Font(name="宋体", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    stats = data.get("match_stats", {})
    info_rows = [
        ("匹配策略", data.get("strategy_name", "")),
        ("银行流水总行数", stats.get("total_bank_rows", 0)),
        ("筛选命中行数", stats.get("filtered_rows", 0)),
        ("匹配机构数", f"{stats.get('matched_institutions', 0)}/{stats.get('total_summary_institutions', 0)}"),
        ("匹配率", f"{stats.get('match_rate', 0):.1f}%"),
        ("差额比例", f"{stats.get('diff_percentage', 0):.1f}%"),
    ]
    for i, (label, val) in enumerate(info_rows, 3):
        ws.cell(row=i, column=1, value=label).font = Font(name="宋体", size=10, bold=True)
        ws.cell(row=i, column=2, value=val).font = normal_font

    # 差异明细表
    start_row = len(info_rows) + 4
    headers = ["机构", "筛选金额", "回款表金额", "差额", "差额比例", "匹配结果"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    diff = data.get("diff_detail", [])
    for ri, d in enumerate(diff[:50], start_row + 1):
        vals = [
            d.get("机构", ""), d.get("筛选金额", 0), d.get("回款表金额", 0),
            d.get("差额", 0), d.get("差额比例", ""), "匹配" if d.get("差额", 1) == 0 else "差异"
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = normal_font
            cell.border = thin_border
            if ci in (2, 3, 4) and isinstance(v, (int, float)):
                cell.number_format = money_fmt

    # 列宽
    widths = [22, 15, 15, 15, 12, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 数值汇总
    sum_row = start_row + len(diff) + 2
    ws.cell(row=sum_row, column=1, value="合计").font = Font(name="宋体", size=10, bold=True)
    ws.cell(row=sum_row, column=2, value=stats.get("total_filtered_amount", 0)).number_format = money_fmt
    ws.cell(row=sum_row, column=3, value=stats.get("total_summary_amount", 0)).number_format = money_fmt
    ws.cell(row=sum_row, column=4, value=stats.get("total_difference", 0)).number_format = money_fmt
    for ci in range(1, 7):
        ws.cell(row=sum_row, column=ci).border = thin_border
        ws.cell(row=sum_row, column=ci).font = Font(name="宋体", size=10, bold=True)

    wb.save(output_path)
    return output_path
