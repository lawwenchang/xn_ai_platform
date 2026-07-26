#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
审计数据匹配引擎 - 处理跨表数据核对、医保回款匹配等场景。
"""
from __future__ import annotations

import json, os, re
from pathlib import Path
from typing import Any, Dict, List, Optional
import functools
import pandas as pd

# 默认无关键词筛选。LLM 根据用户意图动态生成 pattern，无则全量匹配
FALLBACK_PATTERNS = ""

# ═══════════════════════════════════════════════════════════════
# 提取式核对关键词词典（三级供给：词典命中→LLM提案→人工批准入库）
# ═══════════════════════════════════════════════════════════════

EXTRACTION_KEYWORD_DICT = {
    "医保": {
        "patterns": "医保|医疗保险|医保统筹|医保基金|医保结算|医疗补助|社保回款|医疗统筹",
        "columns": ["摘要", "对方客户名称", "附言", "用途"],
        "note": "含社保回款和医疗统筹是因为部分地区医保/社保共用回款账户",
    },
    "社保": {
        "patterns": "社保|社会保险|养老|失业|工伤|生育|社保基金|社保补贴",
        "columns": ["摘要", "对方客户名称", "附言"],
    },
    "公积金": {
        "patterns": "公积金|住房公积金|住房补贴|住房基金",
        "columns": ["摘要", "对方客户名称"],
    },
    "教育经费": {
        "patterns": "教育经费|教育附加|教育费附加|地方教育附加",
        "columns": ["摘要", "附言"],
    },
    "工会经费": {
        "patterns": "工会经费|工会会费|工会",
        "columns": ["摘要", "附言"],
    },
    "政府补助": {
        "patterns": "政府补助|财政拨款|专项资金|补贴|补助款|财政补贴",
        "columns": ["摘要", "对方客户名称", "附言"],
    },
}


def resolve_extraction_keywords(user_intent: str) -> dict:
    """三级供给：词典命中→直接返回；未命中→LLM提案→返回（标注llm_proposed）。
    
    Returns:
        {"patterns": "医保|统筹|社保", "columns": [...], "source": "dict"|"llm"}
    """
    intent_lower = user_intent.lower() if user_intent else ""
    # 1) 词典精确命中
    for key, entry in EXTRACTION_KEYWORD_DICT.items():
        if key in intent_lower:
            return {"patterns": entry["patterns"], "columns": entry["columns"],
                    "source": "dict", "dict_key": key,
                    "note": entry.get("note", "")}
    # 2) LLM 提案
    llm_pattern = _extract_patterns_via_llm(user_intent)
    if llm_pattern:
        return {"patterns": llm_pattern, "columns": ["摘要", "对方客户名称"],
                "source": "llm_proposed",
                "note": "LLM 自动提案，建议人工确认后纳入词典"}
    # 3) 兜底空
    return {"patterns": "", "columns": [], "source": "none",
            "note": "无法确定筛选关键词，将全表参与比对"}


def reverse_validate_unmatched(unmatched_df: pd.DataFrame, desc_col: str = "摘要",
                                top_n: int = 10) -> dict:
    """反向校验：对未匹配行按摘要聚类 top-N，供审计师扫一眼发现误杀。
    
    Returns:
        {"cluster_count": N, "clusters": [{"key": "摘要前6字", "count": N, "sample": "..."}, ...],
         "total_unmatched": N}
    """
    if unmatched_df.empty or desc_col not in unmatched_df.columns:
        return {"cluster_count": 0, "clusters": [], "total_unmatched": len(unmatched_df)}
    from collections import Counter
    keys = unmatched_df[desc_col].fillna("(空摘要)").astype(str).apply(
        lambda x: x.strip()[:8] if x.strip() else "(空摘要)")
    top_clusters = keys.value_counts().head(top_n)
    clusters = []
    for key, count in top_clusters.items():
        sample = unmatched_df[keys == key][desc_col].iloc[0] if count > 0 else ""
        clusters.append({"key": key, "count": count,
                         "sample": str(sample)[:60] if sample is not None else ""})
    return {"cluster_count": len(clusters), "clusters": clusters,
            "total_unmatched": len(unmatched_df)}

@functools.lru_cache(maxsize=256)
def _extract_patterns_via_llm(user_intent: str) -> str:
    """用 LLM 从用户意图中提取筛选关键词。
    同一 intent 全运行期只调一次：保证同任务两次运行筛选口径一致（审计证据可复现）。"""
    try:
        import requests, re
        vllm_url = os.environ.get("VLLM_TUNNEL_URL", "http://localhost:18000/v1/chat/completions")
        prompt = f"审计数据筛选专家。从用户需求中提取筛选关键词（用|分隔），只返回关键词字符串，不要解释。\n\n用户需求：{user_intent}\n\n关键词："
        r = requests.post(
            vllm_url,
            json={"model": "qwen3-235b", "messages": [{"role": "user", "content": prompt}],
                  "temperature": 0.1, "max_tokens": 50},
            headers={"Authorization": "Bearer EMPTY"},
            timeout=15,
        )
        if r.status_code == 200:
            result = r.json()["choices"][0]["message"]["content"].strip()
            cleaned = re.sub(r'[^一-龥a-zA-Z|]', '', result)
            if cleaned:
                return cleaned
    except Exception as e:
        print(f"[LLM] 关键词提取失败: {e}")
    return ""

def detect_file_type(df: pd.DataFrame, filename: str) -> str:
    """智能识别文件类型：journal / bank_statement / summary_table / unknown"""
    cols_str = " ".join(str(c).lower() for c in df.columns)
    fname_lower = filename.lower()
    # 序时账/日记账强特征：凭证号、科目编码、借贷双列+凭证组合
    journal_kw = ["凭证号", "凭证编号", "凭证号码", "科目编码", "科目名称"]
    if any(k in fname_lower for k in ("序时账", "日记账", "明细账")) \
            or sum(1 for k in journal_kw if k in cols_str) >= 1 \
            or ("借方金额" in cols_str and "贷方金额" in cols_str and "凭证" in cols_str):
        return "journal"
    bank_kw = ["交易日期", "摘要", "收入", "支出", "余额", "对方", "借方", "贷方"]
    if "流水" in fname_lower or "bank" in fname_lower or sum(1 for k in bank_kw if k in cols_str) >= 3:
        return "bank_statement"
    summary_kw = ["合计", "汇总", "总计"]
    if "汇总" in fname_lower or "情况表" in fname_lower or sum(1 for k in summary_kw if k in cols_str) >= 2:
        return "summary_table"
    return "unknown"


def identify_columns(df: pd.DataFrame, file_type: str) -> Dict[str, str]:
    """识别数据表的关键列名，返回多个候选列用于多列融合匹配"""
    cols_lower = {str(c).lower(): str(c) for c in df.columns}
    mapping = {}
    if file_type == "bank_statement":
        mapping["date_col"] = _find_col(cols_lower, ["交易日期", "日期", "date", "时间"])
        mapping["desc_col"] = _find_col(cols_lower, ["摘要", "摘要信息", "描述", "说明", "用途", "备注", "description"])
        mapping["counterparty_col"] = _find_col(cols_lower, ["对方客户名称", "对方户名", "对方", "交易对手"])
        mapping["counterparty_org_col"] = _find_col(cols_lower, ["对方组织单元名称", "对方机构"])
        mapping["amount_col"] = _find_col(cols_lower, ["交易金额", "金额", "发生额", "amount"])
        # 方向修正（银行流水第一常识）：收入=贷方（收入），支出=借方（支取）
        mapping["income_col"] = _find_col(cols_lower, ["收入", "收入金额", "贷方（收入）", "贷方(收入)", "贷方金额", "credit", "income"])
        mapping["expense_col"] = _find_col(cols_lower, ["支出", "支出金额", "借方（支取）", "借方(支取)", "借方金额", "debit", "expense"])
        mapping["account_col"] = _find_col(cols_lower, ["银行账号", "账号", "账户", "银行账户"])
        mapping["balance_col"] = _find_col(cols_lower, ["余额", "账户余额", "期末余额"])
    elif file_type == "journal":
        mapping["date_col"] = _find_col(cols_lower, ["日期", "记账日期", "交易日期", "业务日期"])
        mapping["voucher_col"] = _find_col(cols_lower, ["凭证号", "凭证号码", "凭证编号", "凭证字号"])
        mapping["desc_col"] = _find_col(cols_lower, ["摘要", "摘要信息", "说明", "用途", "备注"])
        mapping["debit_col"] = _find_col(cols_lower, ["借方金额", "借方", "借方发生额"])
        mapping["credit_col"] = _find_col(cols_lower, ["贷方金额", "贷方", "贷方发生额"])
        mapping["account_col"] = _find_col(cols_lower, ["银行账号", "账号", "账户"])
        mapping["subject_col"] = _find_col(cols_lower, ["科目编码", "科目名称", "会计科目"])
    elif file_type == "summary_table":
        mapping["name_col"] = _find_col(cols_lower, ["客户名称", "机构名称", "单位", "名称", "机构", "医院名称"])
        mapping["total_col"] = _find_col(cols_lower, ["合计", "总计", "汇总", "sum", "total"])
    # 语义兜底：候选表未命中的关键角色交给统一语义注册表
    try:
        from core.column_semantics import detect_column_roles
        roles = detect_column_roles(df)
        for k, role in {"date_col": "date", "amount_col": "amount", "name_col": "name",
                        "counterparty_col": "counterpart", "desc_col": "summary",
                        "account_col": "account", "balance_col": "balance"}.items():
            if not mapping.get(k) and role in roles:
                mapping[k] = roles[role]
    except Exception:
        pass
    return mapping


def _find_col(cols_lower: dict, candidates: list) -> Optional[str]:
    for c in candidates:
        if c.lower() in cols_lower:
            return cols_lower[c.lower()]
    for c in candidates:
        for cl, orig in cols_lower.items():
            if c.lower() in cl:
                return orig
    return None


def _extract_amount(row, income_col: str, expense_col: str) -> float:
    """从银行流水行提取金额（收入为正，支出为负）"""
    amt = 0.0
    if income_col and income_col in row.index:
        try:
            v = float(row[income_col])
            if not pd.isna(v):
                amt += v
        except (ValueError, TypeError):
            pass
    if expense_col and expense_col in row.index:
        try:
            v = float(row[expense_col])
            if not pd.isna(v):
                amt -= v
        except (ValueError, TypeError):
            pass
    return amt


# 医保场景特化的行政区划清洗词（仅医保匹配链路显式开启，全局禁用——
# 否则"朝阳区甲公司"与"海淀区甲公司"会被误判为同一家，造成张冠李戴）
MEDICAL_STRIP_WORDS = ("市", "县", "区", "中心", "管理")


def _fuzzy_match_name(n1: str, n2: str, strip_admin: bool = False) -> bool:
    """模糊匹配两个机构名称。strip_admin=True 仅用于医保回款场景。"""
    if strip_admin:
        clean = lambda s: s.replace("市", "").replace("县", "").replace("区", "").replace("中心", "").replace("管理", "")
        a, b = clean(n1), clean(n2)
    else:
        a, b = n1.strip(), n2.strip()
    if a == b:
        return True
    if len(a) >= 4 and len(b) >= 4 and a[:4] == b[:4]:
        return True
    if len(a) >= 3 and len(b) >= 3 and (a in b or b in a):
        return True
    return False


def _extract_institution_from_desc(desc: str, known: list) -> Optional[str]:
    """从描述文本中提取机构名称（通用，不限于医保）"""
    for inst in known:
        short = inst
        for _w in ("市", "县", "区", "医疗保险", "基金管理中心", "新型农村合作医疗"):
            short = short.replace(_w, "")  # 医保场景专用清洗（本函数即医保链路）
        if len(short) >= 2 and short in desc:
            return inst
    patterns = [
        r"([\u4e00-\u9fa5]{2,6}(市|县|区)?(公司|中心|医院|局|所|部|站|社|会|院|行))",
        r"([\u4e00-\u9fa5]{2,4}(公司|中心|医院|局|所|部|站|社|会|院|行))",
    ]
    for p in patterns:
        m = re.search(p, desc)
        if m:
            extracted = m.group(0)
            for inst in known:
                if _fuzzy_match_name(extracted, inst):
                    return inst
            return extracted
    return None

def _find_best_match_amount(target_name: str, amounts: dict):
    """在筛选金额中查找最佳匹配。
    返回 (found: bool, amount: float)。净额口径下金额可为 0 或负，
    不能再用 amt>0 当'找到'的判据。"""
    if target_name in amounts:
        return True, amounts[target_name]
    for name, amt in amounts.items():
        if _fuzzy_match_name(target_name, name):
            return True, amt
    return False, 0.0

def match_medical_insurance(bank_df, summary_df, bank_fn="", summary_fn="", patterns="",
                            kw_source: str = "", kw_version: str = ""):
    """医保回款匹配核心：筛选银行流水 → 按机构汇总 → 与回款表比对。
    v3.11 口径修正：净额汇总（退费/冲正自然互抵），退费单独列示——
    回款=贷方净发生额，绝不用 abs() 虚增。"""
    if not patterns:
        patterns = FALLBACK_PATTERNS
    b_cols = identify_columns(bank_df, "bank_statement")
    s_cols = identify_columns(summary_df, "summary_table")
    desc_col = b_cols.get("desc_col")
    counterparty_col = b_cols.get("counterparty_col")
    counterparty_org_col = b_cols.get("counterparty_org_col")
    amount_col = b_cols.get("amount_col")
    income_col = b_cols.get("income_col")
    expense_col = b_cols.get("expense_col")
    name_col = s_cols.get("name_col")

    # 1. 多列联合筛选（含规则级命中统计，不变）
    total_bank = len(bank_df)
    mask = pd.Series([False] * len(bank_df), index=bank_df.index)
    rule_stats = {"filter_columns": [], "keyword_hits": {}, "fallback_full_table": False,
                  "keyword_source": kw_source, "keyword_version": kw_version}
    filter_cols = [c for c in [desc_col, counterparty_col, counterparty_org_col]
                   if c and c in bank_df.columns]
    for col in filter_cols:
        col_mask = bank_df[col].astype(str).str.contains(patterns, na=False, case=False)
        rule_stats["filter_columns"].append({"column": col, "hits": int(col_mask.sum())})
        mask = mask | col_mask
    for kw in [k.strip() for k in str(patterns).split("|") if k.strip()][:50]:
        kw_mask = pd.Series([False] * len(bank_df), index=bank_df.index)
        for col in filter_cols:
            kw_mask = kw_mask | bank_df[col].astype(str).str.contains(
                kw, na=False, case=False, regex=False)
        rule_stats["keyword_hits"][kw] = int(kw_mask.sum())
    if mask.any():
        medical = bank_df[mask].copy()
    else:
        medical = bank_df.copy()
        rule_stats["fallback_full_table"] = True

    # 2. 按机构汇总（净额 + 退费单列）
    known_insts = []
    if name_col and name_col in summary_df.columns:
        known_insts = [str(n).strip() for n in summary_df[name_col].dropna().tolist()
                       if str(n).strip() and str(n).strip() != "合计"]
    inst_amounts, refunds = {}, {}
    for _, row in medical.iterrows():
        desc = str(row.get(desc_col, "")) if desc_col else ""
        cpty = str(row.get(counterparty_col, "")) if counterparty_col else ""
        cpty_org = str(row.get(counterparty_org_col, "")) if counterparty_org_col else ""
        if amount_col and amount_col in row.index:
            try:
                amt = float(row[amount_col])
                if pd.isna(amt):
                    amt = _extract_amount(row, income_col, expense_col)
            except (ValueError, TypeError):
                amt = _extract_amount(row, income_col, expense_col)
        else:
            amt = _extract_amount(row, income_col, expense_col)
        matched = (_extract_institution_from_desc(cpty, known_insts) or
                   _extract_institution_from_desc(cpty_org, known_insts) or
                   _extract_institution_from_desc(desc, known_insts) or
                   (cpty[:30] if cpty else desc[:30]))
        inst_amounts[matched] = inst_amounts.get(matched, 0) + amt      # ← 净额
        if amt < 0:                                                     # ← 退费/冲回单列
            refunds[matched] = refunds.get(matched, 0) + abs(amt)

    match_logic = {
        "筛选列": [c for c in [desc_col, counterparty_col, counterparty_org_col] if c and c in bank_df.columns],
        "筛选模式": patterns,
        "机构识别列": [c for c in [counterparty_col, counterparty_org_col, desc_col] if c and c in bank_df.columns],
        "金额列": amount_col or (f"收入({income_col})-支出({expense_col})" if income_col or expense_col else "自动"),
        "金额口径": "净额（退费互抵），退费单独列示",
    }

    result = _build_match_result(inst_amounts, summary_df, total_bank, len(medical),
                                 b_cols, s_cols, "多列联合匹配（摘要+对方客户名称+对方机构）", match_logic)
    result["rule_stats"] = rule_stats
    result["refund_by_inst"] = {k: round(v, 2) for k, v in refunds.items() if v > 0}
    result["filtered_medical_rows"] = medical
    return result

def match_by_counterparty_only(bank_df, summary_df, bank_fn="", summary_fn="", patterns=""):
    """策略2：仅通过对方客户名称筛选和识别。净额口径，退费单列。"""
    b_cols = identify_columns(bank_df, "bank_statement")
    s_cols = identify_columns(summary_df, "summary_table")
    cpty_col = b_cols.get("counterparty_col")
    amount_col = b_cols.get("amount_col")
    name_col = s_cols.get("name_col")
    patterns = FALLBACK_PATTERNS
    total_bank = len(bank_df)
    if cpty_col and cpty_col in bank_df.columns:
        mask = bank_df[cpty_col].astype(str).str.contains(patterns, na=False)
        medical = bank_df[mask].copy()
    else:
        medical = bank_df.copy()
    known_insts = [str(n).strip() for n in summary_df[name_col].dropna().tolist()
                   if str(n).strip() != "合计"] if name_col in summary_df.columns else []
    inst_amounts, refunds = {}, {}
    for _, row in medical.iterrows():
        cpty = str(row.get(cpty_col, "")) if cpty_col else ""
        try:
            amt = float(row[amount_col]) if amount_col and amount_col in row.index \
                else _extract_amount(row, b_cols.get("income_col"), b_cols.get("expense_col"))
            if pd.isna(amt):
                amt = 0.0
        except (ValueError, TypeError):
            amt = 0.0
        matched = _extract_institution_from_desc(cpty, known_insts) or (cpty[:30] if cpty else "未识别")
        inst_amounts[matched] = inst_amounts.get(matched, 0) + amt      # ← 净额（原 abs 删除）
        if amt < 0:
            refunds[matched] = refunds.get(matched, 0) + abs(amt)
    result = _build_match_result(inst_amounts, summary_df, total_bank, len(medical),
                                 b_cols, s_cols, "对方客户名称匹配",
                                 {"筛选列": [cpty_col], "机构识别": cpty_col, "金额口径": "净额"})
    result["refund_by_inst"] = {k: round(v, 2) for k, v in refunds.items() if v > 0}
    return result

def _build_match_result(inst_amounts, summary_df, total_bank, filtered_rows,
                        b_cols, s_cols, strategy_name, match_logic):
    """构建统一格式的匹配结果。
    v3.11：matched 判定用'机构名是否找到'（found 标志），不再用 amt>0；
    合计口径双侧净额，差异比例用 |分母| 防零。"""
    name_col = s_cols.get("name_col")
    total_col = s_cols.get("total_col")
    summary_insts = {}
    if name_col and total_col and name_col in summary_df.columns and total_col in summary_df.columns:
        for _, row in summary_df.iterrows():
            n = str(row[name_col]).strip()
            if n and n != "合计" and n != "nan":
                try:
                    summary_insts[n] = float(row[total_col])
                except (ValueError, TypeError):
                    summary_insts[n] = 0.0
    diff_list, total_filt, matched_count = [], 0.0, 0
    total_summ = sum(summary_insts.values())                    # ← 净额（原 abs 删除）
    for inst, s_amt in summary_insts.items():
        found, f_amt = _find_best_match_amount(inst, inst_amounts)
        if found:
            matched_count += 1
        total_filt += f_amt
        d = f_amt - s_amt
        dp = abs(d) / abs(s_amt) * 100 if s_amt != 0 else 0
        diff_list.append({"机构": inst, "筛选金额": round(f_amt, 2), "回款表金额": round(s_amt, 2),
                          "差额": round(d, 2), "差额比例": f"{dp:.1f}%"})
    for inst, amt in inst_amounts.items():
        if inst not in summary_insts and inst != "未识别":
            total_filt += amt
            diff_list.append({"机构": inst, "筛选金额": round(amt, 2), "回款表金额": 0,
                              "差额": round(amt, 2), "差额比例": "N/A"})
    total_diff = total_filt - total_summ
    diff_pct = abs(total_diff) / abs(total_summ) * 100 if total_summ != 0 else 100
    mr = matched_count / len(summary_insts) * 100 if summary_insts else 0
    return {"diff_summary": diff_list,
            "match_stats": {"total_bank_rows": total_bank, "filtered_rows": filtered_rows,
                            "total_summary_institutions": len(summary_insts),
                            "matched_institutions": matched_count,
                            "total_filtered_amount": round(total_filt, 2),
                            "total_summary_amount": round(total_summ, 2),
                            "total_difference": round(total_diff, 2),
                            "diff_percentage": round(diff_pct, 2),
                            "match_rate": round(mr, 2), "匹配率": round(mr, 2),
                            "差额比例": round(diff_pct, 2)},
            "bank_cols": b_cols, "summary_cols": s_cols,
            "match_logic": match_logic, "strategy_name": strategy_name}

def _run_all_strategies(bank_df, summary_df, bank_fn="", summary_fn="", patterns="",
                        kw_source: str = "", kw_version: str = ""):
    """策略执行：由列可得性探测选定唯一采用策略（先验判定，非赛后比分），
    其余策略结果留痕备查（alternative_strategies），不自动挑优。"""
    if not patterns:
        patterns = FALLBACK_PATTERNS
    strategies = []
    # 策略1：仅摘要列筛选
    try:
        b_cols = identify_columns(bank_df, "bank_statement")
        s_cols = identify_columns(summary_df, "summary_table")
        desc_col = b_cols.get("desc_col")
        if desc_col and desc_col in bank_df.columns:
            mask = bank_df[desc_col].astype(str).str.contains(patterns, na=False)
            single = bank_df[mask].copy()
            known = [str(n).strip() for n in summary_df[s_cols["name_col"]].dropna().tolist()
                     if str(n).strip() != "合计"] if s_cols.get("name_col") in summary_df.columns else []
            inst_amts, refunds1 = {}, {}
            for _, row in single.iterrows():
                desc = str(row[desc_col])
                amt = _extract_amount(row, b_cols.get("income_col"), b_cols.get("expense_col"))
                m = _extract_institution_from_desc(desc, known) or "未识别"
                inst_amts[m] = inst_amts.get(m, 0) + amt              # ← 净额（原 abs 删除）
                if amt < 0:
                    refunds1[m] = refunds1.get(m, 0) + abs(amt)
            r = _build_match_result(inst_amts, summary_df, len(bank_df), len(single),
                                    b_cols, s_cols, "仅摘要列筛选", {"筛选列": [desc_col], "金额口径": "净额"})
            r["refund_by_inst"] = {k: round(v, 2) for k, v in refunds1.items() if v > 0}
            strategies.append(r)
    except Exception as e:
        print(f"[策略] 策略1失败: {e}")
    # 策略2：对方客户名称列匹配
    try:
        strategies.append(match_by_counterparty_only(bank_df, summary_df, bank_fn, summary_fn))
    except Exception as e:
        print(f"[策略] 策略2失败: {e}")
    # 策略3：多列联合匹配
    try:
        strategies.append(match_medical_insurance(bank_df, summary_df, bank_fn, summary_fn, kw_source=kw_source, kw_version=kw_version))
    except Exception as e:
        print(f"[策略] 策略3失败: {e}")
    if not strategies:
        raise ValueError("所有匹配策略均失败")

    # 探测选定（先验：信息更多的策略优先；不比分）
    b_cols = identify_columns(bank_df, "bank_statement")
    if b_cols.get("counterparty_col") or b_cols.get("counterparty_org_col"):
        adopted = "多列联合匹配（摘要+对方客户名称+对方机构）"
    elif b_cols.get("desc_col"):
        adopted = "仅摘要列筛选"
    else:
        adopted = strategies[0]["strategy_name"]
    best = next((s for s in strategies if s["strategy_name"] == adopted), strategies[0])
    best["alternative_strategies"] = [
        {"strategy": s["strategy_name"],
         "match_rate": s["match_stats"]["match_rate"],
         "diff_pct": s["match_stats"]["diff_percentage"],
         "adopted": False}
        for s in strategies if s is not best]
    best["match_logic"]["策略选定依据"] = (
        "列探测：存在对方户名/机构列→多列联合；仅摘要列→单摘要策略。"
        "其余策略结果留痕备查，未按匹配率挑优。")
    return best, strategies


def run_matching_pipeline(input_dir: Path, output_dir: Path, patterns: str = "",
                          kw_source: str = "", kw_version: str = "") -> dict:
    """
    完整匹配流水线：读取输入 → 识别文件类型 → 执行匹配 → 导出结果
    patterns: 筛选关键词（从 LLM 生成的 DAG 中提取，空则用默认）
    kw_source: 关键词来源（如 "dictionary_v3.0"、"llm_proposed"、"user_approved@2026-07-24"）
    kw_version: 关键词版本
    """
    if not patterns:
        patterns = FALLBACK_PATTERNS
    excel_files = [f for f in input_dir.glob("*")
                   if f.suffix.lower() in (".xlsx", ".xls", ".csv",
                                           ".docx", ".doc", ".pdf", ".md", ".txt")]
    if len(excel_files) < 2:
        raise ValueError(f"需要至少2个文件，当前仅有 {len(excel_files)} 个")

    dfs = {}
    ftypes = {}
    for f in excel_files:
        try:
            if f.suffix.lower() in (".xlsx", ".xls", ".csv"):
                if f.suffix.lower() == ".csv":
                    df = pd.read_csv(f, encoding="utf-8-sig")
                else:
                    df = _read_excel_auto_header(f)
            else:
                # 文档格式（docx/doc/pdf/md/txt）：经统一文档加载器提取表格
                from core.document_loader import load_tables
                tables = load_tables(f)
                if not tables:
                    raise ValueError("文档中未提取到表格")
                df = tables[0]
            dfs[f.name] = df
            ftypes[f.name] = detect_file_type(df, f.name)
        except Exception as e:
            print(f"[警告] 无法读取 {f.name}: {e}")



    # 识别文件角色
    bank_file = summary_file = None
    for name, ft in ftypes.items():
        if ft == "bank_statement" and not bank_file:
            bank_file = name
        elif ft == "summary_table" and not summary_file:
            summary_file = name
    if not bank_file or not summary_file:
        fnames = list(dfs.keys())
        if len(fnames) >= 2:
            bank_file, summary_file = fnames[0], fnames[1]
            if "汇总" in bank_file or "情况表" in bank_file:
                bank_file, summary_file = summary_file, bank_file

    if not bank_file or not summary_file:
        raise ValueError("无法识别银行流水和汇总表文件")

    print(f"[匹配引擎] 银行流水: {bank_file}, 汇总表: {summary_file}")

    # 清理汇总表：去除底部政策说明行
    sdf = dfs[summary_file]
    if not sdf.empty:
        fc = sdf.columns[0]
        found_total = False
        keep = pd.Series([True] * len(sdf))
        for i, (_, row) in enumerate(sdf.iterrows()):
            v = str(row[fc]).strip()
            if v == "合计":
                found_total = True
            elif found_total and v and len(v) > 20:
                keep.iloc[i] = False
        sdf = sdf[keep]
    else:
        sdf = sdf

    # 多策略并行匹配
    result, all_strategies = _run_all_strategies(dfs[bank_file], sdf, bank_file, summary_fn=summary_file, patterns=patterns, kw_source=kw_source, kw_version=kw_version)
    print(f"[匹配引擎] 最优策略: {result['strategy_name']}, 匹配率: {result['match_stats']['match_rate']}%, 差额: {result['match_stats']['diff_percentage']}%")

    # 导出
    output_dir.mkdir(parents=True, exist_ok=True)
    if result["diff_summary"]:
        pd.DataFrame(result["diff_summary"]).to_csv(
            output_dir / "analysis_result.csv", index=False, encoding="utf-8-sig")
    # 多策略对比表
    strategy_comparison = []
    for s in all_strategies:
        ms = s["match_stats"]
        strategy_comparison.append({
            "策略": s["strategy_name"],
            "匹配率": f"{ms['match_rate']:.1f}%",
            "差额": f"{ms['total_difference']:,.2f}",
            "差额比例": f"{ms['diff_percentage']:.1f}%",
            "筛选行数": ms["filtered_rows"],
        })
    if len(strategy_comparison) > 1:
        pd.DataFrame(strategy_comparison).to_csv(
            output_dir / "strategy_comparison.csv", index=False, encoding="utf-8-sig")
    summary_json = {
        "total_rows": len(result["diff_summary"]),
        "columns": list(result["diff_summary"][0].keys()) if result["diff_summary"] else [],
        "match_stats": result["match_stats"],
        "diff_summary": result["diff_summary"],
        "match_logic": result.get("match_logic", {}),
        "strategy_name": result.get("strategy_name", ""),
        "all_strategies": strategy_comparison,
        "numeric_summary": {
            "筛选金额": {"sum": result["match_stats"]["total_filtered_amount"], "mean": 0, "max": 0, "min": 0},
            "回款表金额": {"sum": result["match_stats"]["total_summary_amount"], "mean": 0, "max": 0, "min": 0},
            "差额": {"sum": result["match_stats"]["total_difference"], "mean": 0, "max": 0, "min": 0},
        },
    }
    (output_dir / "journal_entries.json").write_text(json.dumps(summary_json, ensure_ascii=False, indent=2), encoding="utf-8")
    
    # 导出筛选后的银行流水明细（审计师需要看到原始筛选结果）
    if bank_file and bank_file in dfs:
        try:
            b_cols = identify_columns(dfs[bank_file], "bank_statement")
            desc_col = b_cols.get("desc_col")
            cpty_col = b_cols.get("counterparty_col")
            cpty_org_col = b_cols.get("counterparty_org_col")
            mask = pd.Series([False] * len(dfs[bank_file]), index=dfs[bank_file].index)
            for col in [desc_col, cpty_col, cpty_org_col]:
                if col and col in dfs[bank_file].columns:
                    mask = mask | dfs[bank_file][col].astype(str).str.contains(patterns, na=False, case=False)
            if mask.any():
                filtered = dfs[bank_file][mask].copy()
                try:
                    filtered.to_excel(output_dir / "筛选结果_银行流水明细.xlsx", index=False, engine="openpyxl")
                except Exception:
                    filtered.to_csv(output_dir / "筛选结果_银行流水明细.csv", index=False, encoding="utf-8-sig")
                print(f"[匹配引擎] 已导出筛选明细: {len(filtered)} 行")
        except Exception as e:
            print(f"[匹配引擎] 导出筛选明细失败（非致命）: {e}")

    # ── 处置分桶：银端+账端两方未匹配项按六桶归类 ──
    try:
        unmatched_detail = result.get("unmatched_detail", [])
        if unmatched_detail:
            triage_bank = triage_records(unmatched_detail, side="银端未匹配")
            export_triage_board(triage_bank, output_dir,
                                filename="分桶看板_银端.xlsx")
            result["triage_bank"] = triage_bank
            summary_json["triage_summary"] = triage_bank["summary"]
        # 台账端未匹配
        ledger_unmatched = result.get("ledger_unmatched_detail", [])
        if ledger_unmatched:
            triage_ledger = triage_records(ledger_unmatched, side="账端未匹配")
            export_triage_board(triage_ledger, output_dir,
                                filename="分桶看板_账端.xlsx")
            result["triage_ledger"] = triage_ledger
            summary_json["triage_ledger_summary"] = triage_ledger["summary"]
        (output_dir / "journal_entries.json").write_text(
            json.dumps(summary_json, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        print(f"[分桶看板] 生成失败（非致命）: {e}")
    
    print(f"[匹配引擎] 完成。匹配率: {result['match_stats']['match_rate']}%, 差额: {result['match_stats']['diff_percentage']}%")
    return result

def _read_excel_auto_header(filepath):
    """自动检测 Excel 表头行并读取"""
    best_score, best_hr = -1, 0
    for hr in range(6):
        try:
            tmp = pd.read_excel(filepath, header=hr, nrows=0)
            cols = list(tmp.columns)
            score = 0
            for c in cols:
                cs = str(c)
                if cs.startswith("Unnamed"): score -= 1
                elif len(cs) > 2 and not any("\u4e00" <= ch <= "\u9fff" for ch in cs): score -= 1
                else: score += 1
            score += len(cols) * 0.5
            if score > best_score:
                best_score, best_hr = score, hr
        except Exception:
            continue
    if best_score > 0:
        return pd.read_excel(filepath, header=best_hr)
    return pd.read_excel(filepath, header=None)



# ══════════════════════════════════════════════════════
# B5 增强：Blocking 分块 + RapidFuzz 概率评分 + 多对一/一对多 + 置信度路由
# ══════════════════════════════════════════════════════

def _normalize_amount(val) -> float:
    """金额归一化：去千位符、中文逗号 → float；失败返回 0"""
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace(",", "").replace("，", "").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


def _rapidfuzz_weighted_score(r1: dict, r2: dict, weights: dict) -> float:
    """RapidFuzz 加权概率评分。

    r1, r2 是两行记录，weights 定义各字段权重：
    {"amount": 0.35, "date": 0.25, "desc": 0.20, "counterparty": 0.20}
    金额差≤容差→该维满分；日期差越小分数越高；文本用 partial_ratio。
    """
    score = 0.0
    try:
        from rapidfuzz import fuzz
    except ImportError:
        fuzz = None

    if "amount" in weights and weights["amount"] > 0:
        a1 = _normalize_amount(r1.get("amount", 0))
        a2 = _normalize_amount(r2.get("amount", 0))
        if a1 > 0 and a2 > 0:
            ratio = max(0, 1 - min(abs(a1 - a2) / max(a1, a2), 1))
            score += ratio * weights["amount"]

    if "date" in weights and weights["date"] > 0:
        d1, d2 = str(r1.get("date", "")), str(r2.get("date", ""))
        days = 9  # 默认最大容忍差
        try:
            from datetime import datetime
            fmt = "%Y-%m-%d" if "-" in d1 else "%Y/%m/%d"
            days = abs((datetime.strptime(d1[:10], fmt)
                        - datetime.strptime(d2[:10], fmt)).days)
        except Exception:
            pass
        ratio = max(0, 1 - days / 7)
        score += ratio * weights["date"]

    if fuzz and "desc" in weights and weights["desc"] > 0:
        d1, d2 = str(r1.get("desc", "")), str(r2.get("desc", ""))
        if d1 and d2:
            score += (fuzz.partial_ratio(d1, d2) / 100) * weights["desc"]

    if fuzz and "counterparty" in weights and weights["counterparty"] > 0:
        c1 = str(r1.get("counterparty", ""))
        c2 = str(r2.get("counterparty", ""))
        if c1 and c2:
            score += (fuzz.partial_ratio(c1, c2) / 100) * weights["counterparty"]

    return min(1.0, score)

def _block_candidates(bank_rows: list, ledger_rows: list,
                      abs_tol: float = 100.0, date_window_days: int = 3) -> list:
    """Blocking 分块：优先精确金额桶（分）；无精确对应才落入 ±abs_tol（元）窗口。
    禁止按金额百分比开窗；不做静默截断的笛卡尔兜底——候选为空就为空，走人工。"""
    from collections import defaultdict
    exact = defaultdict(list)
    for i, r in enumerate(ledger_rows):
        cents = round(_normalize_amount(r.get("amount", 0)) * 100)
        if cents > 0:
            exact[cents].append((i, r))
    candidates, used_fallback = [], False
    for bi, br in enumerate(bank_rows):
        b_cents = round(_normalize_amount(br.get("amount", 0)) * 100)
        if b_cents <= 0:
            continue
        hit = exact.get(b_cents)
        if hit:
            candidates.extend((bi, li, br, lr) for li, lr in hit)
            continue
        used_fallback = True
        lo, hi = b_cents - abs_tol * 100, b_cents + abs_tol * 100
        for cents, rows_ in exact.items():
            if lo <= cents <= hi:
                candidates.extend((bi, li, br, lr) for li, lr in rows_)
    if used_fallback:
        print(f"[Blocking] 部分记录无精确金额对应，落入±{abs_tol}元窗口桶（需人工关注）")
    return candidates


# 噪音费用词表（仅影响"是否参与逐笔匹配"，不做删除）
# 注意：利息、冲正已从噪音中移出——准则要求关注存款收益与规模匹配性
# （问题解答第12号），冲正重做是典型舞弊手法，二者单独成类输出。
NOISE_FEE_WORDS = ("手续费", "短信费", "年费", "账户管理费", "工本费",
                   "服务费", "测试")
INTEREST_WORDS = ("利息", "结息")
REVERSAL_WORDS = ("冲正", "冲销", "红冲", "撤销")


def _classify_by_wordlist(bank_row: dict, ledger_row: dict) -> str:
    """词表秒判（不消耗 LLM）。命中返回分类，未命中返回空串。"""
    row_text = (json.dumps(bank_row, ensure_ascii=False)
                + json.dumps(ledger_row, ensure_ascii=False))
    if any(w in row_text for w in INTEREST_WORDS):
        return "利息收支（需与存款规模匹配性分析）"
    if any(w in row_text for w in REVERSAL_WORDS):
        return "冲正/重做交易（需关注业务合理性）"
    if any(w in row_text for w in NOISE_FEE_WORDS):
        return "噪音费用"
    return ""


def _propose_cluster_label_via_llm(intent: str, cluster_key: str, count: int) -> str:
    """簇级 LLM 提案：一簇一次调用，产出四选一标签。
    失败/超时返回空串——调用方兜底'待人工核查'。"""
    import os
    try:
        import requests
        vllm_url = os.environ.get("VLLM_TUNNEL_URL",
                                   "http://localhost:18000/v1/chat/completions")
        prompt = (
            f"审计师意图：{intent}\n"
            f"一批（{count}笔）未匹配流水，摘要共性：{cluster_key}\n"
            "分类标准：银行手续费/管理费等无台账对应的小额费用=噪音费用；"
            "跨期到账的时间性差异=未达账项；同一笔业务出现两次=重复入账；"
            "无法解释的金额不符=疑似错报。\n"
            "只输出一个词：未达账项/重复入账/噪音费用/疑似错报")
        r = requests.post(vllm_url,
                          json={"model": "qwen3-235b",
                                "messages": [{"role": "user", "content": prompt}],
                                "temperature": 0, "max_tokens": 10},
                          headers={"Authorization": "Bearer EMPTY"}, timeout=15)
        if r.status_code == 200:
            return r.json()["choices"][0]["message"]["content"].strip()
    except Exception:
        pass
    return ""


def _route_by_confidence(matches: list, intent: str = "") -> dict:
    """置信度三档路由。
    auto ≥0.90 自动确认 / review 0.75-0.89 待人工复核 / exception <0.75 异常分类。
    异常分类：词表秒判 → 摘要聚类 → 每簇一次 LLM 提案（≤10簇）→ 兜底'待人工核查'。
    红线：LLM 只在簇级提案，绝不逐行判定。"""
    routed = {"auto": [], "review": [], "exception": []}
    exceptions = []
    for m in matches:
        score = m.get("confidence", 0)
        if score >= 0.90:
            routed["auto"].append(m)
        elif score >= 0.75:
            m["action"] = "待人工复核"
            routed["review"].append(m)
        else:
            exceptions.append(m)

    rest = []
    for m in exceptions:
        cls = _classify_by_wordlist(m.get("bank_row", {}), m.get("ledger_row", {}))
        if cls:
            m["classification"] = cls
            m["action"] = f"规则判定: {cls}"
        else:
            rest.append(m)
        routed["exception"].append(m)

    from collections import defaultdict
    clusters = defaultdict(list)
    for m in rest:
        key = str(m.get("bank_row", {}).get("desc", ""))[:6] or "(空摘要)"
        clusters[key].append(m)
    for i, (key, members) in enumerate(clusters.items()):
        label = "待人工核查"
        if i < 10:
            label = _propose_cluster_label_via_llm(intent, key, len(members)) or label
        for m in members:
            m["classification"] = label
            m["action"] = f"聚类提案: {label}（{len(members)}笔同类）"
    return routed






# ═══════════════════════════════════════════════════════════════
# 通用分桶入口（所有场景可调）
# ═══════════════════════════════════════════════════════════════

def triage_records(records: list, side: str = "", materiality: float = 500000,
                    small_threshold: float = 10000, **kwargs) -> dict:
    """通用分桶入口，加 side 标签（"银端"/"账端"），所有场景可调。
    内部调 triage_unmatched，结果带回 side 标记。
    """
    result = triage_unmatched(records, materiality=materiality,
                               small_threshold=small_threshold, **kwargs)
    result["side"] = side
    for b in result.get("buckets", []):
        b["side"] = side
    return result

# ═══════════════════════════════════════════════════════════════
# 处置分桶器：把未匹配清单压成带建议程序的桶
# ═══════════════════════════════════════════════════════════════

def triage_unmatched(unmatched: list, materiality: float = 500000,
                      small_threshold: float = 10000, window_days: int = 7,
                      period_end_date: str = None) -> dict:
    """六桶优先级分桶：大额→同户重复→同户聚合→摘要同构→期末窗口→小额长尾。
    排完六个桶后剩余的行才叫"待人工核查"。
    """
    from collections import defaultdict, Counter
    rows = list(unmatched)
    total = len(rows)
    if total == 0:
        return {"buckets": [], "remaining": [], "summary": {"total": 0, "bucketed": 0, "remaining": 0, "bucketed_pct": 0}}
    assigned = set()

    def _amt(r):
        for k in ("金额", "交易金额", "发生额", "net_amount", "amount"):
            v = r.get(k, 0)
            if v:
                try: return abs(float(v))
                except (ValueError, TypeError): pass
        return 0.0

    def _name(r):
        for k in ("对方客户名称", "对手方", "counterpart", "机构名称", "单位"):
            v = r.get(k, "")
            if v and str(v).strip(): return str(v).strip()
        return ""

    def _desc(r):
        for k in ("摘要", "desc", "summary"):
            v = r.get(k, "")
            if v and str(v).strip(): return str(v).strip()
        return ""

    def _date(r):
        for k in ("交易日期", "date", "日期"):
            v = r.get(k, "")
            if v:
                try: return pd.to_datetime(v)
                except Exception: pass
        return None

    def _make_bucket(name, idxs, proc, sample_n=3):
        items = [rows[i] for i in idxs]
        amt = sum(_amt(r) for r in items)
        for i in idxs: assigned.add(i)
        return {"name": name, "rows": items, "count": len(items),
                "amount": round(amt, 2), "procedure": proc, "sample": items[:sample_n]}

    print(f"[分桶] 未匹配 {total} 笔，开始六桶优先级分桶...")

    buckets = []
    # 桶1：大额
    large = [i for i, r in enumerate(rows) if i not in assigned and _amt(r) >= materiality]
    if large:
        buckets.append(_make_bucket("大额簇（≥50万）", large, "逐笔抽凭全查"))
        amt_sum = sum(_amt(rows[i]) for i in large)
        print(f"[分桶] 大额簇: {len(large)}笔 {amt_sum:,.0f}元 (≥{materiality/10000:.0f}万)")

    # 桶2：同户名同额重复≥3笔
    key_ct = Counter(); key_idx = defaultdict(list)
    for i, r in enumerate(rows):
        if i in assigned: continue
        key_ct[(_name(r), round(_amt(r), 2))] += 1
        key_idx[(_name(r), round(_amt(r), 2))].append(i)
    repeat = []
    for (n, a), c in key_ct.items():
        if c >= 3: repeat.extend(key_idx[(n, a)])
    if repeat:
        buckets.append(_make_bucket("同户名同额重复（≥3笔）", repeat, "抽1笔核凭证推断全簇"))
        amt_sum = sum(_amt(rows[i]) for i in repeat)
        print(f"[分桶] 同户名同额重复: {len(repeat)}笔 {amt_sum:,.0f}元")

    # 桶3：同户名聚合
    name_idx = defaultdict(list)
    for i, r in enumerate(rows):
        if i in assigned: continue
        n = _name(r)
        if n: name_idx[n].append(i)
    for n, idxs in sorted(name_idx.items(), key=lambda x: len(x[1]), reverse=True)[:8]:
        if len(idxs) >= 2:
            buckets.append(_make_bucket(f"同户名聚合: {n[:12]}", idxs, f"对账单/函证核对总额（{len(idxs)}笔）不逐笔"))
    agg_total = sum(b["count"] for b in buckets if "同户名聚合" in b["name"])
    if agg_total:
        print(f"[分桶] 同户名聚合: {agg_total}笔（{sum(1 for b in buckets if '同户名聚合' in b['name'])}个对手方）")

    # 桶4：摘要同构≥10笔
    desc_idx = defaultdict(list)
    for i, r in enumerate(rows):
        if i in assigned: continue
        d = _desc(r)[:6]
        if d: desc_idx[d].append(i)
    desc_clustered = 0
    for p, idxs in sorted(desc_idx.items(), key=lambda x: len(x[1]), reverse=True):
        if len(idxs) >= 10:
            buckets.append(_make_bucket(f"摘要同构: {p}...", idxs, f"定性后检查代表性样本推断全簇"))
            desc_clustered += len(idxs)
    if desc_clustered:
        print(f"[分桶] 摘要同构: {desc_clustered}笔")

    # 桶5：期末窗口
    if period_end_date:
        try:
            pend = pd.to_datetime(period_end_date)
            win = [i for i, r in enumerate(rows) if i not in assigned
                   and _date(r) and abs((_date(r) - pend).days) <= window_days]
            if win:
                buckets.append(_make_bucket(f"期末窗口±{window_days}天", win, "期后验证未达候选"))
                print(f"[分桶] 期末窗口±{window_days}天: {len(win)}笔")
        except Exception: pass

    # 桶6：小额长尾
    small = [i for i, r in enumerate(rows) if i not in assigned and 0 < _amt(r) < small_threshold]
    if small:
        buckets.append(_make_bucket("小额长尾（<1万）", small, "分析程序月度总额波动合理即过"))
        amt_sum = sum(_amt(rows[i]) for i in small)
        print(f"[分桶] 小额长尾: {len(small)}笔 {amt_sum:,.0f}元 (<{small_threshold/10000:.0f}万)")

    # 剩余
    remaining = [rows[i] for i in range(total) if i not in assigned]
    if remaining:
        print(f"[分桶] 待人工核查: {len(remaining)}笔 ({len(remaining)/max(total,1)*100:.1f}%)")
    else:
        print(f"[分桶] 全部归类完毕，0笔待人工核查")
    return {"buckets": buckets, "remaining": remaining,
            "summary": {"total": total, "bucketed": total - len(remaining),
                        "remaining": len(remaining),
                        "bucketed_pct": round((total - len(remaining)) / max(total, 1) * 100, 1)}}



def export_triage_board(triage_result: dict, output_dir: Path,
                         filename: str = "分桶看板.xlsx") -> str:
    """导出分桶看板 Excel（每桶一个 sheet，剩余一个 sheet）。"""
    output_dir.mkdir(parents=True, exist_ok=True)
    board_path = output_dir / filename
    with pd.ExcelWriter(board_path, engine="openpyxl") as writer:
        # 汇总 sheet - 锚点+活公式
        from core.formula_writer import export_triage_summary_with_formulas
        export_triage_summary_with_formulas(writer.book, triage_result)
        # 每桶一个 sheet
        for i, b in enumerate(triage_result.get("buckets", [])):
            df = pd.DataFrame(b["rows"])
            if not df.empty:
                safe_name = re.sub(r'[\/*?:\[\]<>]', '_', b['name'][:15])
                df.to_excel(writer, sheet_name=f"桶{i+1}_{safe_name}", index=False)
        # 剩余
        rem = triage_result.get("remaining", [])
        if rem:
            pd.DataFrame(rem).to_excel(writer, sheet_name="待人工核查", index=False)
    print(f"[分桶看板] {board_path} ({len(triage_result.get('buckets',[]))}桶/"
          f"{triage_result['summary']['total']}笔→{triage_result['summary']['remaining']}笔待人工核查)")
    side_label = triage_result.get("side", "").replace("未匹配", "")
    for b in triage_result.get("buckets", []):
        if b["count"] > 50:
            try:
                safe_name2 = re.sub(r'[\/*?:\[\]<>]', '_', b['name'][:20])
                wp_name = f"核查底稿_{side_label}_{safe_name2}.xlsx"
                export_counterpart_workpaper(b, output_dir, filename=wp_name)
            except Exception as _wp_err:
                print(f"[分桶看板] 核查底稿生成失败: {_wp_err}")
    return str(board_path)


# ═══════════════════════════════════════════════════════════════
# 核查指引模板（三层拼装：方向→旗型→特征叠加）
# ═══════════════════════════════════════════════════════════════

GUIDE_TEMPLATES = {
    "整月单边": {
        "流入": [
            "1.调取企业全部银行账户流水，核对该笔是否记在他行科目（排除良性串户）",
            "2.查银行回单用途/附言，确认资金性质（借款/货款/还款）",
            "3.若为账外收入，追查对应业务合同与发票，评估漏记范围",
            "4.访谈财务：该月该户资金流入未记账的原因"],
        "流出": [
            "1.调取他行流水，排除良性串户（收款方是否自家账户）",
            "2.查付款回单+审批单，确认资金去向与授权",
            "3.追踪收款账户后续流向（关联方/个人/回流本企业他户）",
            "4.若为账外支出，评估体外循环/挪用风险，考虑舞弊风险升级"],
    },
    "整数大额": {
        "流入": [
            "1.查资金来源性质：对方户名→回单用途→借款/注资/还款协议",
            "2.过桥测试：期后30天内是否有等额或近似等额流出",
            "3.核对关联方清单，判断是否关联方资金往来",
            "4.若为经营收款，函证或核对销售合同/发票"],
        "流出": [
            "1.查资金去向：回单+审批单，确认付款对象与授权",
            "2.过桥测试：期后30天内是否有等额或近似等额流入（回流）",
            "3.核对关联方清单，判断是否关联方拆借",
            "4.若为经营付款，核对采购合同/发票/入库单"],
    },
    "一收一付同额": {
        "流入": [
            "1.比对收付两笔的对方户名：同户=疑似过桥/刷流水；异户=核对两笔业务凭证",
            "2.查两笔的时间间隔与回单用途",
            "3.核实是否融资性过桥（期末冲存款规模嫌疑）"],
        "流出": [
            "1.同流入1：比对对方户名是否同户",
            "2.查付款与收款的先后逻辑，判断是否资金空转",
            "3.关注是否虚增交易量配合虚开发票"],
    },
    "分次转入转出": {
        "流入": ["1.合并查看同户多笔流入总额，判断是否拆分规避审批/监管",
                "2.查后续是否有集中转出，资金是否过手性质"],
        "流出": ["1.合并查看同户多笔流出总额，判断是否拆分付款规避审批权限",
                "2.核对付款审批单的单笔限额与实际拆分情况"],
    },
    "费用小额月度差异": {
        "流出": ["1.核对企业全部银行账户，确认该费用是否从他户扣款",
                "2.索取银行收费回单，比对账面计提与实扣差异原因",
                "3.连续多月固定差额→费用扣款账户可能未提供，要求补充流水"],
    },
    "疑似重复入账": {
        "双向": ["1.调两笔凭证比对附件（回单号是否相同）",
                "2.确认是否同一业务重复记账，建议冲销分录"],
    },
}


def build_guide(flag_type, net, same_day_same_amount=False,
                period_end_window=False, round_amount=False):
    """三层模板拼装：旗型定主模板→方向定变体→特征叠加定升级提示。"""
    direction = "流入" if net > 0 else "流出"
    tpl = GUIDE_TEMPLATES.get(flag_type, {})
    steps = tpl.get(direction) or tpl.get("双向") or ["1.人工分析该笔业务背景与凭证"]
    mods = []
    if same_day_same_amount:
        mods.append("⚡同日复向等额：优先过桥/空转测试")
    if period_end_window:
        mods.append("⚡期末窗口：叠加粉饰嫌疑，优先级上调")
    if round_amount and flag_type != "整数大额":
        mods.append("⚡整数金额：注意非经营性资金性质")
    return "\n".join(mods + steps)



# ═══════════════════════════════════════════════════════════════
# 大额簇核查底稿：按对手方分组 + 预计算检查列 + 规律摘要
# ═══════════════════════════════════════════════════════════════

def export_counterpart_workpaper(bucket, output_dir, filename="核查底稿.xlsx"):
    """大额簇按对手方分组生成核查底稿。"""
    from collections import Counter, defaultdict
    rows = bucket.get("rows", [])
    if not rows: return ""
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / filename
    def _g(r,*ks):
        for k in ks:
            v=r.get(k)
            if v is not None and v!="": return v
        return ""
    def _amt(r):
        for k in ("net_amount","金额","amount","净额"):
            v=r.get(k,0)
            if v:
                try: return float(v)
                except: pass
        return 0.0
    def _date(r):
        for k in ("date","日期","交易日期"):
            v=r.get(k,"")
            if v:
                try: return pd.to_datetime(v)
                except: pass
        return pd.NaT
    def _cp(r): return _g(r,"counterpart","对手方","对方客户名称","对方","机构名称")
    def _desc(r): return _g(r,"summary","摘要","desc")

    dates=[_date(r) for r in rows]
    valid=[d for d in dates if pd.notna(d)]
    max_date=max(valid) if valid else pd.NaT

    cp_groups=defaultdict(list); no_cp=[]
    for r in rows:
        cp=_cp(r).strip()
        if cp: cp_groups[cp].append(r)
        else: no_cp.append(r)
    for cp in cp_groups: cp_groups[cp].sort(key=lambda r:str(_date(r)))

    date_amt_map=Counter()
    for r in rows:
        d=_date(r); a=round(_amt(r),2)
        if pd.notna(d) and a>0: date_amt_map[(d.date().isoformat(),a)]+=1

    cp_monthly=Counter()
    for r in rows:
        cp=_cp(r).strip(); d=_date(r)
        if cp and pd.notna(d): cp_monthly[(cp,d.strftime("%Y-%m"))]+=1

    total_amt=sum(abs(_amt(r)) for r in rows)
    cp_stats=[(cp,len(g),sum(abs(_amt(r)) for r in g)) for cp,g in cp_groups.items()]
    cp_stats.sort(key=lambda x:x[2],reverse=True)
    top_n=min(6,len(cp_stats)); top_cps=cp_stats[:top_n]
    top_amt=sum(s[2] for s in top_cps)
    top_pct=round(top_amt/max(total_amt,1)*100,1)

    summary_lines=[
        "桶名: "+bucket.get("name",""),
        "总笔数: %d  |  总金额: %s元"%(len(rows),total_amt),
        "对手方数: %d  |  无对手方: %d笔"%(len(cp_groups),len(no_cp)),
        "前%d大户: 占%.1f%%金额（%s元）"%(top_n,top_pct,top_amt),
    ]
    for cp,cnt,amt in top_cps: summary_lines.append("  · %s: %d笔 %s元"%(cp[:20],cnt,amt))


    work_rows = []
    sorted_cps = sorted(cp_groups.keys(), key=lambda x: sum(abs(_amt(r)) for r in cp_groups[x]), reverse=True)
    month_str = max_date.strftime("%Y-%m") if pd.notna(max_date) else ""
    for cp in sorted_cps:
        group = cp_groups[cp]
        group_amt = sum(abs(_amt(r)) for r in group)
        is_red_flag = cp_monthly.get((cp, month_str), 0) >= 5 if month_str else False
        work_rows.append({
            "对手方": "▼ " + cp + "（%d笔，%s元）" % (len(group), group_amt),
            "红旗户": "⚠ 整月单边" if is_red_flag else "",
            "日期": "", "摘要": "", "净额": "", "金额整数": "", "同日同额": "", "期末窗口": "",
            "分类": "", "核查指引": "",
        })
        for r in group:
            d = _date(r); a = _amt(r)
            d_str = d.date().isoformat() if pd.notna(d) else ""
            if abs(a) == int(abs(a)) and abs(a) > 0: is_int = "✓ 整数"
            else: is_int = "✗ 非整数"
            if pd.notna(d) and date_amt_map.get((d.date().isoformat(), round(a, 2)), 0) >= 2: same_day = "✓ 同日同额"
            else: same_day = "✗ 无同日同额"
            if pd.notna(d) and pd.notna(max_date) and (max_date - d).days <= 7: near_end = "✓ 期末±7天"
            else: near_end = "✗ 距期末>7天"

            orig_cls = _g(r, "classification", "分类")
            cls_parts = []
            if is_red_flag: cls_parts.append("整月单边")
            if abs(a) == int(abs(a)) and abs(a) >= 100000: cls_parts.append("整数大额")
            if orig_cls and orig_cls not in ("待人工核查", "待核查", ""): cls_parts.append(orig_cls)
            if not cls_parts: cls_parts.append("待核查")
            classification = " + ".join(cls_parts)

            same_day_flag = pd.notna(d) and date_amt_map.get((d.date().isoformat(), round(a, 2)), 0) >= 2
            near_end_flag = pd.notna(d) and pd.notna(max_date) and (max_date - d).days <= 7
            amt_round = abs(a) == int(abs(a)) and abs(a) > 0
            guide_type = "整月单边" if is_red_flag else ("整数大额" if (amt_round and abs(a) >= 100000) else "")
            audit_guide = build_guide(guide_type, a, same_day_same_amount=same_day_flag,
                                      period_end_window=near_end_flag, round_amount=amt_round)

            work_rows.append({
                "对手方": cp, "红旗户": "⚠ 整月单边" if is_red_flag else "",
                "日期": d_str, "摘要": _desc(r)[:60], "净额": a,
                "金额整数": is_int, "同日同额": same_day, "期末窗口": near_end,
                "分类": classification, "核查指引": audit_guide,
            })

    work_df = pd.DataFrame(work_rows)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame({"规律摘要": summary_lines}).to_excel(writer, sheet_name="规律摘要", index=False)
        work_df.to_excel(writer, sheet_name="核查底稿", index=False)
        ws = writer.sheets["核查底稿"]
        from openpyxl.styles import Font, PatternFill
        yf = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        rf = Font(color="FF0000", bold=True); bf = Font(bold=True); gf = Font(color="008000", bold=True)
        for row_idx in range(1, len(work_rows) + 1):
            rd = work_rows[row_idx - 1]; er = row_idx + 1
            if str(rd.get("对手方", "")).startswith("▼"):
                for ci in range(1, len(work_df.columns) + 1):
                    ws.cell(row=er, column=ci).font = bf
                    ws.cell(row=er, column=ci).fill = yf
            if str(rd.get("红旗户", "")).startswith("⚠"):
                ws.cell(row=er, column=2).font = rf
            for cn in ["金额整数", "同日同额", "期末窗口"]:
                if rd.get(cn) and rd.get(cn)[0] == "✓":
                    ci = list(work_df.columns).index(cn) + 1
                    ws.cell(row=er, column=ci).font = gf
        cw = {"对手方": 24, "红旗户": 14, "日期": 12, "摘要": 42, "净额": 14,
              "金额整数": 16, "同日同额": 16, "期末窗口": 16, "分类": 22, "核查指引": 48}
        for cn, w in cw.items():
            if cn in work_df.columns:
                cl = chr(65 + list(work_df.columns).index(cn))
                ws.column_dimensions[cl].width = w
    print("[核查底稿] %s (%d个对手方/%d笔 前%d大户占%.0f%%)" % (path, len(cp_groups), len(rows), top_n, top_pct))
    return str(path)



# 命令行入口
if __name__ == "__main__":
    import sys
    idir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("inputs")
    odir = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("outputs")
    try:
        run_matching_pipeline(idir, odir)
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)