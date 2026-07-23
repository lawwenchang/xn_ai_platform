#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
纯中文 Excel/Pandas Q&A 生成器 (zh_excel_qa_bank.py)
=====================================================
用途：第三轮微调补训数据——修复第二轮训练集中 Excel 领域中英文混搭问题。
特点：
- 全部 system/user/assistant 均为中文
- 公式答案由模板确定性生成，保证正确
- 参数融入审计场景（银行流水/交易金额/摘要/科目编码），贴合事务所语境
"""
import json
import random

SYSTEM = "你是数据处理与Excel专家，精通Excel公式、Pandas和数据分析。准确回答用户的技术问题，给出可直接使用的公式或代码。"


def gen_all():
    items = []
    seen = set()

    def add(q, a):
        if q not in seen:
            seen.add(q)
            items.append((q, a))

    # ═══ 1. IF 单条件 ═══
    for c in ["A2", "B2", "C2", "D2"]:
        for t, hi, lo in [(60, "及格", "不及格"), (90, "优秀", "一般"), (100, "达标", "未达标"),
                          (500000, "大额", "正常"), (0, "正数", "非正数"),
                          (10000, "超标", "正常"), (200, "合格", "不合格"), (5000, "需复核", "通过"),
                          (1000000, "特别重大", "一般"), (300, "高", "低"),
                          (50000, "重点关注", "常规"), (80, "通过", "退回")]:
            f = f'=IF({c}>{t},"{hi}","{lo}")'
            add(f"如果{c}单元格的值大于{t}就显示“{hi}”，否则显示“{lo}”，公式怎么写？",
                f"使用 IF 函数：{f}。IF 的语法是 IF(条件, 条件成立时的值, 不成立时的值)。当 {c} 大于 {t} 时返回“{hi}”，否则返回“{lo}”。如需包含等于 {t} 的情况，把 > 改为 >=。")
            add(f"帮我写个公式：{c}大于{t}返回{hi}，不然返回{lo}",
                f"公式：{f}。IF 函数第一个参数是判断条件（{c}>{t}），第二个参数是条件成立时的返回值，第三个是不成立时的返回值。文本必须用英文双引号包裹。")

    # ═══ 2. 嵌套 IF 分档 ═══
    for c in ["B2", "C2"]:
        for b in [(90, 80, 70, 60), (85, 75, 65, 55)]:
            f = f'=IF({c}>={b[0]},"A",IF({c}>={b[1]},"B",IF({c}>={b[2]},"C",IF({c}>={b[3]},"D","F"))))'
            add(f"分数在{c}，想按 {b[0]} 以上是A、{b[1]} 以上是B、{b[2]} 以上是C、{b[3]} 以上是D、其余是F 来评级，怎么写？",
                f"用嵌套 IF 从高到低依次判断：{f}。注意条件必须按从高到低的顺序排列，否则低档条件会先命中导致结果错误。Excel 2019 及以上也可以用 IFS 函数，可读性更好。")
            add(f"帮我按{b[0]}/{b[1]}/{b[2]}/{b[3]}四个分界点给{c}的成绩分成A到F五档",
                f"嵌套 IF 写法：{f}。每个 IF 的“否则”分支里再嵌套下一个 IF。也可用 IFS：=IFS({c}>={b[0]},\"A\",{c}>={b[1]},\"B\",{c}>={b[2]},\"C\",{c}>={b[3]},\"D\",TRUE,\"F\")，最后的 TRUE 作为兜底。")

    # ═══ 3. IF + AND / OR ═══
    for c1, c2 in [("A2", "B2"), ("C2", "D2")]:
        add(f"{c1}大于500000并且{c2}等于“已审核”时返回“重点关注”，否则返回“正常”，怎么写？",
            f'=IF(AND({c1}>500000,{c2}="已审核"),"重点关注","正常")。AND 函数要求所有条件同时成立才返回 TRUE。如果只需任一条件成立，改用 OR。')
        add(f"只要{c1}是“医保”或者{c2}是“统筹”，就标记为“医保相关”，公式怎么写？",
            f'=IF(OR({c1}="医保",{c2}="统筹"),"医保相关","")。OR 函数任一条件成立即返回 TRUE。若需匹配包含关系（而非完全相等），可以用 =IF(OR(ISNUMBER(FIND("医保",{c1})),ISNUMBER(FIND("统筹",{c2}))),"医保相关","")。')

    # ═══ 4. SUMIF / SUMIFS ═══
    for cat in ["医保回款", "办公费", "差旅费", "手续费", "水电费",
                "利息收入", "房租", "培训费", "维修费", "保险费", "广告费", "咨询费"]:
        for cc, sc in [("A", "C"), ("B", "D"), ("A", "B")]:
            add(f"{cc}列是类别，{sc}列是金额，我想汇总所有“{cat}”的金额合计",
                f'=SUMIF({cc}:{cc},"{cat}",{sc}:{sc})。SUMIF 的三个参数依次是：条件区域、条件、求和区域。这会把 {cc} 列等于“{cat}”的对应 {sc} 列金额全部加总。')
            add(f"帮我算一下{cc}列等于“{cat}”时{sc}列的合计数",
                f'用 SUMIF：=SUMIF({cc}:{cc},"{cat}",{sc}:{sc})。如果需要模糊匹配（摘要中包含“{cat}”即算），改用通配符：=SUMIF({cc}:{cc},"*{cat}*",{sc}:{sc})。')
            add(f"表里{cc}列是科目、{sc}列是发生额，把“{cat}”的发生额加总一下",
                f'=SUMIF({cc}:{cc},"{cat}",{sc}:{sc})。注意：如果“{cat}”在数据里带有空格或全角字符，SUMIF 会匹配不上，先用 TRIM 清洗或改用通配符 "*{cat}*"。')

    for cat in ["医保回款", "销售收入", "采购付款", "往来款", "工程款", "备用金", "押金", "税费"]:
        add(f"想汇总A列为“{cat}”且B列日期在2024年1月1日之后的C列金额，怎么写？",
            f'=SUMIFS(C:C,A:A,"{cat}",B:B,">2024/1/1")。SUMIFS 的第一个参数是求和区域，后面成对出现“条件区域,条件”。日期条件要写成带比较符的文本，如 ">2024/1/1"。')
        add(f"多条件求和：类别是“{cat}”、金额大于10万的记录合计（类别在A列、金额在C列）",
            f'=SUMIFS(C:C,A:A,"{cat}",C:C,">100000")。SUMIFS 支持对同一列同时作为条件列和求和列。所有条件为“且”的关系，必须同时满足。')

    # ═══ 5. COUNTIF / COUNTIFS ═══
    for c in ["A", "B", "C", "D"]:
        for op, v, desc in [(">", 50000, "大于5万"), (">=", 100, "大于等于100"), ("<", 0, "为负数"), (">", 1000000, "超过100万")]:
            add(f"统计{c}列中{desc}的单元格有多少个",
                f'=COUNTIF({c}:{c},"{op}{v}")。COUNTIF 第一个参数是统计区域，第二个是条件。比较条件要用英文双引号包成文本，如 "{op}{v}"。')

    for c1, v1, c2, v2 in [("A", "医保", "C", ">100000"), ("B", "已审核", "D", ">0"),
                           ("A", "北区", "C", ">=50000"), ("B", "未达账", "D", "<0")]:
        add(f"同时满足{c1}列等于“{v1}”且{c2}列{v2}的行数怎么统计？",
            f'=COUNTIFS({c1}:{c1},"{v1}",{c2}:{c2},"{v2}")。COUNTIFS 支持多组“区域,条件”，全部条件同时满足才计数。')

    # ═══ 6. AVERAGEIF / AVERAGEIFS ═══
    for c, cond, desc in [("B", ">0", "正数"), ("C", ">10000", "大于1万"), ("D", "<>0", "非零")]:
        add(f"求{c}列中{desc}的平均值",
            f'=AVERAGEIF({c}:{c},"{cond}")。AVERAGEIF 只对满足条件的单元格求平均。注意：如果没有任何单元格满足条件会返回 #DIV/0! 错误，可用 IFERROR 包裹处理。')
    for cat in ["IT部", "财务部", "审计一部"]:
        add(f"A列是部门，C列是金额，求“{cat}”的平均金额",
            f'=AVERAGEIF(A:A,"{cat}",C:C)。三参数用法：条件区域、条件、求平均的区域。多条件时用 AVERAGEIFS，语法与 SUMIFS 类似（求值区域放第一个参数）。')

    # ═══ 7. VLOOKUP / XLOOKUP / INDEX+MATCH ═══
    for kc, rc, n in [("A", "B", 2), ("A", "C", 3), ("B", "D", 3)]:
        add(f"根据{kc}列的编码查找对应第{n}列的值（表区域 {kc}2:{rc}100），比如查“P007”",
            f'=VLOOKUP("P007",{kc}2:{rc}100,{n},FALSE)。四个参数：查找值、表区域、返回第几列、FALSE 表示精确匹配。注意查找值必须位于表区域的第一列，否则查不到。')
        add(f"VLOOKUP查{kc}2:{rc}100找不到值时显示#N/A，怎么让它显示“未找到”？",
            f'用 IFERROR 包裹：=IFERROR(VLOOKUP(查找值,{kc}2:{rc}100,{n},FALSE),"未找到")。IFERROR 会捕获所有错误并返回第二个参数。只想捕获 #N/A 用 IFNA 更精确。')

    add("科目编码在C列、科目名称在A列，想根据编码反查名称（编码列在名称列右边），VLOOKUP做不到怎么办？",
        '=INDEX(A:A,MATCH(查找编码,C:C,0))。VLOOKUP 只能从左往右查，INDEX+MATCH 无方向限制：MATCH 在 C 列定位行号，INDEX 从 A 列按行号取值。Excel 365 也可以用 =XLOOKUP(查找编码,C:C,A:A)，更直观。')
    add("XLOOKUP和VLOOKUP有什么区别？我该用哪个？",
        'XLOOKUP 是 VLOOKUP 的全面升级：1.支持从右往左查；2.默认精确匹配（VLOOKUP 默认模糊匹配易出错）；3.自带“找不到时的返回值”参数，不用套 IFERROR；4.可以返回多列。语法：=XLOOKUP(查找值,查找列,返回列,"未找到")。注意 XLOOKUP 需要 Excel 365 或 2021，旧版本只能用 VLOOKUP 或 INDEX+MATCH。')
    add("两个条件的查找：根据“机构名称”和“期间”两列同时匹配取回款金额，怎么写？",
        '方法一（Excel 365）：=XLOOKUP(1,(A:A=目标机构)*(B:B=目标期间),C:C,"未找到")。方法二（通用）：=INDEX(C:C,MATCH(1,(A:A=目标机构)*(B:B=目标期间),0))，旧版 Excel 需按 Ctrl+Shift+Enter 作为数组公式输入。方法三：加辅助列 =A2&"|"&B2 拼接双键后用普通 VLOOKUP。')
    add("银行流水的对方户名和台账的机构名称不完全一致（一个有“有限公司”后缀），怎么模糊匹配？",
        '用通配符 VLOOKUP：=VLOOKUP(A2&"*",台账!A:B,2,FALSE)，在查找值后拼 * 可匹配以 A2 开头的机构全称。更稳妥的做法是先做名称清洗：=SUBSTITUTE(SUBSTITUTE(A2,"有限公司",""),"股份","")，两边都去掉常见后缀再精确匹配。大批量模糊匹配建议用 Pandas 处理。')

    # ═══ 8. 日期函数 ═══
    add("怎么获取今天的日期？要能自动更新的",
        '=TODAY() 返回当前日期（每次打开或重算文件自动更新）；=NOW() 返回日期+时间。想要固定不变的当前日期，用快捷键 Ctrl+;（分号）录入静态值。')
    for c in ["A2", "B2"]:
        add(f"计算{c}的日期距离今天有多少天",
            f'=TODAY()-{c}。日期在 Excel 中本质是数字，直接相减即可得到天数。如果结果显示成了日期格式，把单元格格式改为“常规”或“数值”。')
        add(f"{c}是开始日期，B列是结束日期，只算工作日天数（去掉周末），怎么写？",
            f'=NETWORKDAYS({c},B2)。NETWORKDAYS 自动排除周六周日。要再排除法定节假日，把假日列表放在某个区域（如 F2:F20），写成 =NETWORKDAYS({c},B2,F2:F20)。注意中国的调休补班日无法自动识别，需手工调整假日区域。')
    add("从交易日期中提取年份和月份，比如把2024/3/15变成“2024年3月”",
        '=YEAR(A2)&"年"&MONTH(A2)&"月"，或者用 TEXT 函数：=TEXT(A2,"yyyy年m月")。TEXT 更简洁且可控制格式（mm 会补零显示 03）。按月汇总银行流水时常用 =TEXT(A2,"yyyy-mm") 作为辅助列再配合 SUMIF。')
    add("怎么求某个月的最后一天？比如求交易日期A2所在月份的月末日",
        '=EOMONTH(A2,0)。EOMONTH 第二个参数是月份偏移：0 为当月月末，-1 为上月月末，1 为下月月末。审计中做截止测试时，=EOMONTH(A2,0)=A2 可以判断某笔交易是否恰好发生在月末最后一天。')
    add("判断交易日期是不是周末（可能是异常入账）",
        '=IF(WEEKDAY(A2,2)>=6,"周末","工作日")。WEEKDAY(A2,2) 返回 1（周一）到 7（周日），大于等于 6 即周六或周日。筛查费用报销、收入入账中的周末记账是常见的舞弊筛查手段。')
    add("两个日期之间隔了几个月怎么算？",
        '=DATEDIF(A2,B2,"m") 返回完整月数。DATEDIF 的第三个参数：\"d\"=天数、\"m\"=整月数、\"y\"=整年数、\"ym\"=不满一年的月数。注意 DATEDIF 是隐藏函数，输入时没有参数提示，但功能正常。')

    # ═══ 9. 文本函数 ═══
    add("从摘要里判断是否包含“医保”两个字，包含就标记出来",
        '=IF(ISNUMBER(FIND("医保",A2)),"医保相关","")。FIND 返回关键词位置，找不到时返回错误，所以用 ISNUMBER 判断。FIND 区分大小写、SEARCH 不区分且支持通配符。批量多关键词筛选（医保|统筹|社保）建议用 Pandas 的 str.contains。')
    add("银行账号中间部分打码：只显示前4位和后4位，中间用星号",
        '=LEFT(A2,4)&"****"&RIGHT(A2,4)。LEFT 取左边 4 位、RIGHT 取右边 4 位，中间拼接星号。这是底稿对外提供时账号脱敏的常用做法。')
    add("单元格前后有空格导致VLOOKUP匹配不上，怎么清理？",
        '=TRIM(A2) 去除前后空格及词间多余空格。如果还有不可见字符（从系统导出常见），叠加 CLEAN：=TRIM(CLEAN(A2))。全角空格 TRIM 处理不了，需要 =SUBSTITUTE(A2,"　","")（引号里是全角空格）。清洗后再做匹配。')
    add("把A列的姓和B列的名合并成一个单元格",
        '=A2&B2 直接拼接；中间要加分隔符就写 =A2&" "&B2。多列批量合并用 TEXTJOIN：=TEXTJOIN("-",TRUE,A2:D2)，第二个参数 TRUE 表示跳过空单元格（需 Excel 2016+）。')
    add("从“凭证号-2024-0315”这种格式里提取中间的年份",
        '=MID(A2,FIND("-",A2)+1,4)。FIND 定位第一个“-”的位置，MID 从其后 1 位开始取 4 个字符。若格式固定也可以直接 =MID(A2,5,4)。复杂拆分建议用“数据→分列”按“-”分隔符拆开。')
    add("金额数字怎么显示成带千分位、保留两位小数的格式？比如1250000变成1,250,000.00",
        '=TEXT(A2,"#,##0.00")。TEXT 把数字转成指定格式的文本，符合审计底稿的金额书写规范。注意转换后是文本不能再参与求和，仅用于展示；如果只是显示需要，更好的做法是设置单元格格式为“数值→使用千位分隔符→2位小数”，不改变数值本身。')

    # ═══ 10. 重复值 / 核对 ═══
    add("怎么找出A列里的重复凭证号并标记出来？",
        '辅助列输入 =IF(COUNTIF(A:A,A2)>1,"重复","") 并下拉。COUNTIF 统计每个值在整列出现的次数，大于 1 即重复。只标记第二次及以后出现的：=IF(COUNTIF(A$2:A2,A2)>1,"重复","")（注意区域起点锁定）。也可用条件格式→突出显示重复值直接标色。')
    add("两张表核对：找出A列（本表）里有、但C列（对方表）里没有的记录",
        '=IF(COUNTIF(C:C,A2)=0,"对方缺失","")。COUNTIF 在 C 列中数 A2 出现的次数，等于 0 说明对方表没有这条记录。双向核对时再反向做一次 =IF(COUNTIF(A:A,C2)=0,"本表缺失","")。这是未达账项排查的基本方法。')
    add("删除一列数据里的重复值，只保留唯一值",
        '方法一：选中数据→“数据”选项卡→“删除重复值”（会改动原数据，先备份）。方法二（Excel 365）：=UNIQUE(A2:A100) 在新位置生成去重列表，不动原数据。方法三：数据透视表把该字段拖入行区域，天然去重。')
    add("金额对不上，想快速算两列的差异：D列=B列账面数-C列核对数，并且只显示有差异的",
        'D2 输入 =B2-C2 下拉得到差异额；筛选 D 列不等于 0 即只看有差异的行。考虑浮点尾差时用 =ROUND(B2-C2,2)，两位小数后仍不为零才算真差异。差异行建议再加一列备注差异原因（未达账/跨期/需核实）。')

    # ═══ 11. Pandas 审计场景 ═══
    for fn, col, kw in [("银行流水.xlsx", "摘要", "医保|统筹|社保"),
                        ("费用台账.xlsx", "费用说明", "咨询费|服务费|劳务"),
                        ("明细账.xlsx", "摘要", "预提|暂估|年终奖")]:
        add(f"用Pandas读取{fn}，筛选{col}中包含“{kw.replace('|','、')}”的记录",
            f'import pandas as pd\n\ndf = pd.read_excel("{fn}")\nkeywords = "{kw}"\nfiltered = df[df["{col}"].str.contains(keywords, na=False)]\nprint(f"筛选到 {{len(filtered)}} 条记录")\nfiltered.head()\n\nstr.contains 支持正则表达式（用 | 分隔多个关键词）。na=False 会把缺失值视为不匹配而不是报错。')

    for fn, col, amt in [("银行流水.xlsx", "交易金额", 500000), ("费用台账.xlsx", "金额", 100000),
                         ("科目余额表.xlsx", "期末余额", 1000000)]:
        add(f"Pandas读取{fn}，筛选{col}大于{amt:,}的异常大额交易",
            f'import pandas as pd\n\ndf = pd.read_excel("{fn}")\nthreshold = {amt}\nbig = df[df["{col}"] > threshold].copy()\nbig = big.sort_values("{col}", ascending=False)\nprint(f"{{len(big)}} 笔 {{col}} 超过 {{threshold:,}} 元")\nbig[["{col}"]].describe()\n\ndescribe() 输出 count/mean/std/min/25%/50%/75%/max，快速查看大额交易的统计画像。')

    add("Pandas按月份分组汇总交易金额（日期在“交易日期”列、金额在“交易金额”列）",
        'import pandas as pd\n\ndf = pd.read_excel("银行流水.xlsx")\ndf["年月"] = pd.to_datetime(df["交易日期"]).dt.to_period("M")\nmonthly = df.groupby("年月")["交易金额"].agg(["count","sum","mean"])\nprint(monthly)\n\ngroupby 的 agg 可以一次性输出多个聚合指标。.to_period("M") 将日期标准化为月份（如 2024-01），避免不同日期的同月分属不同组。')
    add("Pandas按科目编码汇总明细账的发生额并合并到科目余额表",
        'import pandas as pd\n\nbal = pd.read_excel("科目余额表.xlsx")\ndetail = pd.read_excel("明细账.xlsx")\ndetail_sum = detail.groupby("科目编码", as_index=False)["发生额"].sum()\nmerged = bal.merge(detail_sum, on="科目编码", how="left")\nmerged["差异"] = merged["余额"] - merged["发生额"].fillna(0)\nprint(merged[merged["差异"].abs() > 0.01])\n\nmerge 类似 SQL JOIN，on 参数指定匹配键。how="left" 保留余额表所有科目，没有明细账的科目发生额为 NaN，用 fillna(0) 填 0。差异阈值 0.01 消除浮点误差。')
    add("Pandas将银行流水和台账按机构名称做outer join，找出双方独有和差异记录",
        'import pandas as pd\n\nliushui = pd.read_excel("银行流水.xlsx")\ntaizhang = pd.read_excel("台账.xlsx")\nmerged = liushui.merge(taizhang, left_on="对方户名", right_on="机构名称",\n                       how="outer", indicator=True)\n# indicator=True 会在结果中增加 _merge 列，标记 left_only/right_only/both\nonly_left = merged[merged["_merge"] == "left_only"]\nonly_right = merged[merged["_merge"] == "right_only"]\nprint(f"仅流水有{{len(only_left)}}条，仅台账有{{len(only_right)}}条")\n\nouter join 保留两边全部记录，是审计核对的标准操作。inner join 只看匹配成功的。')
    add("Pandas删除dataframe里的重复行",
        'df = df.drop_duplicates()  # 删除完全相同的行\n# 按指定列去重\n df = df.drop_duplicates(subset=["凭证号","科目编码"], keep="first")\n# keep="first" 保留第一次出现的记录，"last" 保留最后一次。')
    add("Pandas把计算结果导出为Excel文件，带千分位格式",
        'with pd.ExcelWriter("核对结果.xlsx", engine="openpyxl") as writer:\n    df.to_excel(writer, sheet_name="差异明细", index=False)\n    # 千分位格式通过 openpyxl 设置\n    from openpyxl.styles import numbers\n    ws = writer.sheets["差异明细"]\n    for cell in ws.iter_rows(min_col=2, max_col=3, min_row=2,\n                              max_row=ws.max_row):\n        for c in cell:\n            if isinstance(c.value, (int, float)):\n                c.number_format = \'#,##0.00\'\n')
    add("Pandas读取的CSV文件中文乱码怎么办？",
        'df = pd.read_csv("银行流水.csv", encoding="gbk")     # 常见于国标GBK编码\n# 或者用 utf-8-sig 处理带BOM的UTF-8文件\n# df = pd.read_csv("银行流水.csv", encoding="utf-8-sig")\n# 不确定编码时，先 open(f, "rb") 读二进制看前几个字节判断BOM。')
    add("Pandas读取Excel时，日期列变成了数字怎么办？",
        'df = pd.read_excel("银行流水.xlsx", dtype={"交易日期": str})\n# 然后用 pd.to_datetime 统一解析：\ndf["交易日期"] = pd.to_datetime(df["交易日期"], errors="coerce")\n# errors="coerce" 把无法解析的值转为 NaT（Not a Time），不中断整个过程。')
    add("Pandas处理缺失值：把金额列的空白填为0",
        'df["交易金额"] = df["交易金额"].fillna(0)\n# 如果是一列条件填另一列：\ndf["差异说明"] = df["差异说明"].fillna("待核实")\n# 整表查看缺失情况：\ndf.isnull().sum()  # 每列缺失计数')
    add("Pandas一列想取绝对值怎么做？",
        'df["交易金额"] = df["交易金额"].abs()\n# 或者新生成一列：df["金额绝对值"] = df["交易金额"].abs()\n# 筛选借贷方向不正确的记录（如收入科目出现贷方负数）时可判断\n# df[(df["交易金额"].abs() > 0) & (df["借贷方向"] != "贷")]')
    add("Pandas怎么把字符串类型的列转成数字？",
        'df["交易金额"] = pd.to_numeric(df["交易金额"], errors="coerce")\n# errors="coerce" 把非数字值（如"N/A"）转为 NaN，不报错。\n# 如果整列都是数字但类型是 object，可以用 astype：\ndf["交易金额"] = df["交易金额"].astype(float)')
    add("Pandas按条件标记行：交易金额大于50万标记为“大额”",
        "import numpy as np\n\ndf['标志'] = np.where(df['交易金额'] > 500000, '大额', '正常')\n# np.where 类似 Excel 的 IF 函数，向量化处理整列。\n# 多条件用 np.select：\nconditions = [df['交易金额'] > 1000000, df['交易金额'] > 500000]\nchoices = ['特别大额', '大额']\ndf['级别'] = np.select(conditions, choices, default='正常')")

    # ═══ 12. 扩展补充 ═══
    # --- 取整/舍入 ---
    add("金额四舍五入到两位小数怎么写？",
        "=ROUND(A2,2) 四舍五入到2位小数。审计中用于消除浮点计算尾差。=ROUNDUP(A2,2) 向上取整；=ROUNDDOWN(A2,2) 向下截断。")
    add("ROUND和ROUNDUP的区别？", "ROUND按四舍五入规则舍入；ROUNDUP无论大小一律向上进位。审计中客户计息用ROUNDUP，内部核算用ROUND。")
    add("怎么找一列中的最大值和第N大值？",
        "=MAX(B2:B100) 返回最大值；=LARGE(B2:B100,3) 取第3大；=SMALL(B2:B100,5) 取第5小。审计筛查大额用排序更直观。")
    add("金额排名怎么做？", "=RANK.EQ(C2,C$2:C$100) 返回降序排名（最大排第一）。并列值会跳号。=RANK.AVG 返回平均排名。")
    # --- 数据透视表 ---
    add("数据透视表怎么按月汇总交易金额？", '选中数据→插入→数据透视表→行标签拖入日期→右键组合→按月分组→值区域拖入交易金额。加筛选维度（如对方名称）拖入筛选区域。')
    add("数据透视表怎么显示百分比而不是绝对值？", '右键值区域→值字段设置→值显示方式→"列汇总的百分比"或"总计的百分比"。审计用途：各科目占总资产比重分析。')
    # --- 条件格式 ---
    add("怎么自动把超过50万的行标红？", '选中数据区域→开始→条件格式→突出显示单元格规则→大于→输入500000→红色填充。多规则叠加时注意顺序，排上面的优先。')
    add("金额列用数据条做可视化怎么看？", "选中列→条件格式→数据条。快速扫一眼异常值分布。审计辅助定性用，正式底稿建议保留数值。")
    add("条件格式怎么用公式做灵活判断？", '新建规则→"使用公式确定要设置格式的单元格"→如=AND($B2>500000,$D2="未审核")跨列组合条件。注意绝对列引用$B加相对行引用2。')
    # --- 错误处理 ---
    add("公式出现#N/A是什么意思？", "#N/A表示查找函数未找到匹配值。检查是否有前后空格、数据类型是否一致（文本vs数字）、是否用了精确匹配。可用IFERROR包裹避免影响后续计算。")
    add("DIV/0!错误怎么处理？", '分母为零或空白时出现。改为=IF(B2=0,"",A2/B2)先判断分母。AGGREGATE函数=AGGREGATE(9,6,A2:A10)可自动跳过错误值做求和。')
    add("公式报错但看不出原因怎么办？", '用"公式"选项卡→"公式求值"→逐步查看嵌套结果。或选中公式子表达式按F9查看计算结果（Esc退出不保存）。大规模debug建辅助列拆解。')
    # --- 绝对与相对引用 ---
    add("VLOOKUP下拉时区域跟着跑缩小了查找范围怎么办？", '把表区域改成绝对引用：=VLOOKUP(C2,$A$2:$B$100,2,FALSE)。编辑栏中按F4快速切换引用模式。审计模板中固定参数一律用绝对引用。')
    add("什么时候用A$1、什么时候用$A1？", "A$1只锁行号（下拉不变右拉变列）；$A1只锁列。条件格式经典公式=COUNTIF($A$2:$A2,$A2)>1中混合引用配合才能正确标记重复项。")
    # --- 快捷键与操作 ---
    add("Excel最实用的快捷键有哪些？", "Ctrl+;录入静态日期；Ctrl+Shift+L切换筛选；Ctrl+T创建表；F4切换引用/重复上步；Ctrl+PageDown/Up切换工作表；Alt+=自动求和；Ctrl+1格式设置；F2编辑单元格。")
    add("冻结窗格怎么用？", '要冻结第1行标题和A列索引→选中B2→视图→冻结窗格。审计阅底稿必须冻结表头，否则滚动后就不知道每列是什么了。')
    add("数据分列怎么操作？", '选中列→数据→分列→分隔符号（逗号/空格/\"-\"）→完成。常用于拆分摘要字段。分列会覆盖右侧单元格，先插入空列再操作。')
    add("怎么批量删除空行？", "选中区域→Ctrl+G→定位条件→空值→右键删除整行。或筛选器中找空白值批量删。导入数据的标配第一步清洗。")
    add("Ctrl+T创建表有什么好处？", "公式自动填充新行；列标题自动变筛选器；结构化引用如=SUM(表_流水[金额])替代固定区间。数据透视表以此表为源可自动扩展。")
    add("怎么设置打印区域和页眉页脚？", '页面布局→打印区域→设置打印区域。页眉写"××会计师事务所 审计底稿"，页脚写"第 &[页码] 页/共 &[总页数] 页"。打印前务必设置缩放确保一页宽度。')
    add("大量数据时Excel卡顿怎么办？", "用Ctrl+T建表而非全列引用；去掉不必要条件格式和数组公式；减少INDIRECT/RAND等易失函数；禁用硬件图形加速；大流水超50万行走Pandas。")
    # --- 排序/筛选 ---
    add("怎么按多列排序？", '数据→排序→添加条件：主关键字日期升序，次关键字金额降序。排序前务必加一列递增序号以便恢复原始顺序。排序时勾选"数据包含标题"。')
    add("筛选器里怎么做复合条件？", "文本筛选→自定义筛选→包含/等于/开头可组合且与或。审计筛查时先取消全选再单独勾关注项。Ctrl+Click多选下拉列表值。")
    add("高级筛选和普通筛选的区别？", "高级筛选支持复杂条件组合且结果可复制到其他位置不覆盖原数据。条件区域同行=AND,不同行=OR。普通筛选在万行以上比逐个勾选快。")
    # --- 动态数组 ---
    add("FILTER函数怎么用？", '=FILTER(A2:C100,B2:B100="医保","无数据").FILTER筛出所有匹配行并自动溢出整区域。Excel 365专有，可嵌套在SUM等其他函数里。')
    add("SORT和SORTBY怎么用？", '=SORT(A2:B20,2,-1)按第2列降序；=SORTBY(A2:B20,B2:B20,-1)按B列排序A:B区域。审计用=SORT(FILTER(流水,条件),3,-1)一步筛后排序。')
    add("LET函数怎么用？", '=LET(x,SUM(A2:A100),y,SUM(B2:B100),x/y)。给中间结果取变量名避免重复计算同一子表达式。长公式用LET可读性和性能双提升（Excel 365专有）。')
    # --- SUBTOTAL ---
    add("SUM和SUBTOTAL的区别？", "SUBTOTAL(9,A2:A100)等价SUM但忽略被筛选/隐藏的行。9=SUM,1=AVERAGE,2=COUNT。筛选后只统计可见项用SUBTOTAL。审计小计必备。")
    # --- openpyxl ---
    add("Python的openpyxl怎么设千分位和宋体？",
        "from openpyxl import load_workbook; from openpyxl.styles import Font\nwb=load_workbook('底稿.xlsx'); ws=wb.active\nfor row in ws.iter_rows(min_row=2):\n    for c in row:\n        if isinstance(c.value,(int,float)):\n            c.number_format='#,##0.00'; c.font=Font(name='宋体',size=11)\nwb.save('底稿_格式化.xlsx')")
    add("openpyxl往单元格里写公式怎么写？", 'ws["G2"]="=IF(F2>500000,\\"大额\\",\\"正常\\")"。写入公式时不计算（openpyxl无公式引擎），打开Excel时自动算。批量写公式先用pandas处理导出，openpyxl只做格式微调。')
    # ── 参数化补充 ──
    for kw, col in [("医保", "B"), ("统筹", "D"), ("社保", "A"), ("手续费", "C"),
                    ("罚款", "B"), ("利息", "B"), ("押金", "C"), ("保证金", "A")]:
        add(f"统计{col}列中包含“{kw}”的行数",
            f'=COUNTIF({col}:{col},"*{kw}*")。通配符*匹配任意字符。多关键词叠加用SUMPRODUCT或COUNTIF相加。')
    for d in [3, 5, 7, 10, 15, 20, 30]:
        for c in ["A2", "B2", "TODAY()"]:
            add(f"从{c}起算{d}个工作日之后是哪天？", f'=WORKDAY({c},{d})。WORKDAY跳过周末自动加{d}个工作日。银行回款到账日计算常用。要排除节假日加第三个参数。')
    for p in [5, 10, 20, 25]:
        add(f"找出金额排名前{p}%的异常大额怎么筛选？",
            f'先算>={p}%分位：=PERCENTILE.EXC(C:C,{1-p/100})，再筛选大于该值的行。审计舞弊筛查经典手段。')
    add("PERCENTILE和QUARTILE怎么用？", "=PERCENTILE(C:C,0.25)=Q1第一四分位数；=QUARTILE(C:C,1)同理。审计分析费用报销额分布常用四分位数法判断异常散点。")
    add("怎么计算环比增长率？", '=(B2-A2)/ABS(A2)并设置百分比格式。分母用ABS防负数导致符号混乱。基期为零时用 IF(A2=0,"",(B2-A2)/ABS(A2))兜底。')
    add("怎么按左右方向截取字符？", "=LEFT(A2,N)取左边N字符；=RIGHT(A2,N)取右边N字符；=MID(A2,M,N)从第M位起取N字符。系统导出凭证号/账号拆分常用。")
    add("怎么自动计算行列号？", "=ROW(A2)返回当前行号；=COLUMN(B1)返回列号。INDEX/MATCH中动态调整引用偏移。")
    add("TEXTJOIN怎么拼接多段文本并用分隔符且跳过空单元格？", '=TEXTJOIN(",",TRUE,A2:D2)。第二个参数TRUE跳过空值。Excel2016+可用。')
    add("VLOOKUP不想手动数列号怎么动态算返回列序号？", '=VLOOKUP(A2,表区域,COLUMN(B1)-COLUMN(A1)+1,FALSE)。COLUMN差即为返回列序号，省去在一堆列里数第几列的麻烦。')
    add("怎么把行转置为列？", "=TRANSPOSE(A2:D2)行转列（旧版按Ctrl+Shift+Enter）。Excel365直接溢出。审计偶用于底稿重构：科目月度数据行转列做时序图。")
    add("怎么标记C列金额连续3个月下跌的月份？", '辅助列=IF(AND(C2<C1,C3<C2,C4<C3),"连续下跌"&ROW(C2)&"月起点","")。连续N月逻辑用AND堆叠。审计收入/利润趋势分析用。')
    add("Excel怎么做敏感性分析（变量变化对利润的影响）？", '数据→模拟分析→模拟运算表。列/行变量输入单元格引用，自动算各参数的公式结果。常见于收入变动5%/10%/15%对利润的敏感性测算。多变量同时变化用方案管理器或Solver。')
    add("怎么把多个列的内容合并为一列并带分隔符？", "=A2&B2&C2直接用&拼接。中间加分隔符写成 =A2&\";\"&B2&\";\"&C2。最后复制粘贴值再删中间列。TEXTJOIN做这个更简洁不冗长。")

    # ── 更多参数化变体追击500 ──
    for cat in ["咨询费", "会议费", "招待费", "交通费", "办公用品", "物业费", "取暖费", "邮电费", "运输费"]:
        add(f"A列是费用类别、B列是金额，汇总“{cat}”的总金额和平均金额和笔数",
            f'=SUMIF(A:A,"{cat}",B:B) 得到总金额；=AVERAGEIF(A:A,"{cat}",B:B) 得到平均；=COUNTIF(A:A,"{cat}") 得到笔数。审计费用筛查的标配三件套。')
        add(f"{cat}类的费用有没有单笔超过5万的异常？怎么标记？",
            f'=IF(AND(A2="{cat}",B2>50000),"异常大额","")下拉。按类别筛选后检查业务合理性。或直接用条件格式：=AND($A2="{cat}",$B2>50000)标红。')
    for col in ["C", "D", "E", "F"]:
        for t in [100, 500, 1000, 5000, 10000, 50000]:
            add(f"怎么算{col}列大于{t}的值的平均值？",
                f'=AVERAGEIF({col}:{col},">{t}")。AVERAGEIF单条件求平均，条件内写">="或">"等比较符必须用英文双引号包裹成文本。')
            add(f"{col}列大于{t}的值一共有多少个？", f'=COUNTIF({col}:{col},">{t}")。COUNTIF的行数统计比SUMIF更轻量。')
    for c1, c2 in [("A", "B"), ("B", "C"), ("C", "D")]:
        add(f"如果{c1}2大于{c2}2就显示“对”，否则显示“错”，公式怎么写？",
            f'=IF({c1}2>{c2}2,"对","错")。IF两列比较时引号里改成实际含义（如"已结清"/"待跟进"），用作核对标记。')
    for cc in ["A", "B", "C"]:
        add(f"{cc}列是科目编码，帮我把编码一样的数据按发生额汇总",
            f'用数据透视表把{cc}拖入行标签、发生额拖入值区域。或用 SUMIF({cc}:{cc},{cc}2,发生额列) 做单科目汇总。多科目汇总用数据透视表更快。')
        add(f"{cc}列怎么去掉重复项只留每种编码一条？",
            f'选中{cc}列→数据→删除重复值→只勾{cc}列→确定。或用公式=UNIQUE({cc}:{cc})在新列生成去重列表（Excel 365）。')
    add("怎么计算两个日期之间的年数（精确到小数）？", '=(B2-A2)/365.25。除以365.25考虑闰年均值，比除以365更精确。DATEDIF(A2,B2,"m")/12得到整月转年。')
    add("怎么把数字金额转换成中文大写（壹贰叁）？", "Excel无内置函数，需自定义VBA函数或使用=TEXT(A2,\"[DBNum2]General\")勉强转换但不符财务规范。可靠方案是在模板中预置中文大写金额字符串再用公式拼接。建议审计报告用Word宏做这个。")
    add("同一张表多个条件区域同时求和怎么最快？", "用SUMIFS一次性搞定：=SUMIFS(求和列,条件列1,条件1,条件列2,条件2,...)。不要用多个SUMIF相加——公式长、重复扫描数据、易出错。")
    add("用IF+ISBLANK判断空单元格和IF=\"\"判断有什么不同？", 'ISBLANK(A2)只有真正空单元格返回TRUE；而A2=""对值为空文本（如公式返回\"\"）的也判为TRUE。大多数场景用=""更安全（覆盖实际空值和空文本）。但=IF(A2<>"","有数据","缺数据")对数值且为零的单元格也会判为"有数据"。')
    add("怎么用EXACT函数做区分大小写的精确匹配？", '=IF(EXACT(A2,B2),"完全相同","不同")。EXACT 对大小写和全半角都敏感，=\"\"只比较不区分。审计编码/凭证号比对建议用EXACT防大小写出错。')
    add("怎么把一个单元格里的多行文本（Alt+Enter换行）拆成多行？", '选中单元格→数据→分列→分隔符号→勾选"其他"并输入Ctrl+J（换行符）→完成。Ctrl+J在输入框中不可见但Excel能识别。')
    add("怎么把公式计算的结果固定为值（不再随源表变动）？", '选中公式结果→Ctrl+C→右键→选择性粘贴→值→确定。热键：Ctrl+C→Alt+E→S→V→Enter。审计定稿前务必把依赖外部源表的公式全部贴值固化。')
    add("怎么用数据验证（数据有效性）做输入限制？", '选中单元格→数据→数据验证→允许选择：整数/小数/序列/日期等→设置范围→输入信息和出错警告可选。审计录入模板必备：限制金额必须正数、日期必须在本年度等。')
    add("怎么用INDIRECT根据下拉菜单动态切换引用的工作表？", '=INDIRECT(A1&"!B2")。A1下拉选择工作表名，INDIRECT拼接成动态引用。注意INDIRECT是易失函数（每次改动都重算），大表中慎用。')
    add("怎么用HYPERLINK生成可点击跳转到指定单元格的链接？", '=HYPERLINK("#Sheet2!A1","点此跳转")。审计底稿导航目录常用：索引页上每个科目一行，点击直接跳到对应工作表。第二个参数是显示文本。')
    add("怎么用ISERROR或ISNA检测VLOOKUP结果是否报错？", "=IF(ISNA(VLOOKUP(...)),\"查无此值\",VLOOKUP(...))。ISNA只捕捉#N/A（找不到），ISERROR捕捉所有错误。对两表核对场景，ISNA判断\"对方表里没这条\"是最干净的写法。")
    for amt in [100000, 200000, 300000, 500000, 1000000]:
        add(f"筛选金额大于{amt:,}的交易有多少笔、合计多少金额、平均每笔多少？",
            f'笔数=COUNTIF(C:C,">{amt}")；合计=SUMIF(C:C,">{amt}")；平均=SUMIF(C:C,">{amt}")/COUNTIF(C:C,">{amt}")。审计快速画像大额交易分布。')
    for m in [1, 3, 6, 12]:
        for col in ["A2", "B2"]:
            add(f"从{col}往未来推{m}个月是哪天？",
                f'=EDATE({col},{m})。EDATE加减月数自动处理月末=28~31天的差异。审计中预付费用摊销截止日计算、质保金到期日计算常用。若算{m}个月前，写成=EDATE({col},-{m})。')
            add(f"从{col}所在月份往前/往后{m}个月的月初和月末分别是什么日期？",
                f'月初：=EOMONTH({col},-{m})+1；月末：=EOMONTH({col},0)。EOMONTH({col},N)的N为负数即向前推N个月的月末，+1得到下个月的第一天。')
    for d in [3, 7, 15]:
        add(f"两个日期相差{d}天以内视为同一笔交易（近似匹配），怎么判断？",
            f'=IF(ABS(A2-B2)<={d},\"匹配\",\"不匹配\")。取绝对值ABS比较日期差。核对应收应付时对到账日与记账日之间有{d}天差的情况做容差匹配。')
    add("怎么算一个季度有多少个工作日？",
        '=NETWORKDAYS(start,end,holidays)。年审按季度拆分工作量时常算每季度工作天数。先算出季度首尾日期（如Q1的1/1和3/31），再NETWORKDAYS即可。')
    add("怎么快速知道一个日期是周几（返回中文星期几）？", '=TEXT(A2,"aaaa")返回"星期一"~"星期日"；=TEXT(A2,"ddd")返回"Mon"~"Sun"。审计筛查周末入账：=IF(OR(TEXT(A2,"aaaa")="星期六",TEXT(A2,"aaaa")="星期日"),"周末","工作日")。')

    return items


# ═══════════════ 主程序 ═══════════════
if __name__ == "__main__":
    items = gen_all()
    print(f"生成唯一 Q&A: {len(items)} 条")
    random.seed(42)
    random.shuffle(items)
    wanted = items[:500]  # 取 500 条

    out_path = "d:/Liu/ai_platform_code/phase2_v3/data/finetune/zh_excel_qa_500.jsonl"
    with open(out_path, "w", encoding="utf-8") as f:
        for q, a in wanted:
            rec = {"messages": [
                {"role": "system", "content": SYSTEM},
                {"role": "user", "content": q},
                {"role": "assistant", "content": a},
            ]}
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    print(f"已写入: {out_path} ({len(wanted)} 条)")

    # 验证：确保没有英文疑问句（函数名 VLOOKUP/Pandas 等专有名词属正常）
    import re
    eng = [q for q, _ in wanted if re.match(r'^(How|What|Can|Why|Is|I\s)', q)]
    print(f"英文疑问句开头的记录: {len(eng)}/{len(wanted)}（应为 0）")

